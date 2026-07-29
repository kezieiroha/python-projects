"""How to Run sheet — generated documentation embedded in the workbook.

Intentionally brief. Detailed operational notes belong in README.md; this sheet
exists so the workbook is self-explanatory when opened away from the repo.
"""

from openpyxl.styles import Alignment, Font, PatternFill

SHEET_TITLE = "How to Run"
COMMAND = "python3 build_alert_levels.py"


def render(ws):
    """Write the How to Run sheet."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 90

    def line(row, text, bold=False, size=11, color="000000", bg=None, indent=0):
        c = ws.cell(row=row, column=1, value=(" " * indent) + text)
        c.font = Font(name="Calibri", size=size, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[row].height = 22 if bold else 18
        return row + 1

    r = 1
    r = line(r, "How to Run This Spreadsheet", bold=True, size=14,
             color="1F4E79", bg="DEEAF1")
    r = line(r, "")
    r = line(r, "WHAT IT DOES", bold=True, size=11, color="2E75B6")
    r = line(r, "Generates a timestamped .xlsx in this folder. Current prices, moving "
                "averages, fundamentals, and analyst data are refreshed from Yahoo "
                "Finance. Historical Fib anchors and thesis notes are loaded from "
                "workbook/data/portfolio_rows.json.", size=10, indent=2)
    r = line(r, "")
    r = line(r, "REQUIREMENTS", bold=True, size=11, color="2E75B6")
    r = line(r, "Python 3", size=10, indent=2)
    r = line(r, "openpyxl, yfinance, pandas", size=10, indent=2)
    r = line(r, "")
    r = line(r, "ONE-TIME SETUP  (only needed once)", bold=True, size=11, color="2E75B6")
    r = line(r, "pip install -r requirements.txt", size=10, indent=2, color="7030A0")
    r = line(r, "")
    r = line(r, "RUNNING THE SCRIPT", bold=True, size=11, color="2E75B6")
    r = line(r, "Run from the investment-tracker folder:", size=10, indent=2)
    r = line(r, "")

    c = ws.cell(row=r, column=1, value=f"    {COMMAND}")
    c.font = Font(name="Courier New", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 24
    r += 1

    r = line(r, "")
    r = line(r, "Output pattern:", size=10, indent=2)
    r = line(r, "SIPP_Alert_Levels_v21_DDMMYY-HHMM.xlsx", size=10, indent=4, color="7030A0")
    r = line(r, "")
    r = line(r, "LIVE DATA REFRESH", bold=True, size=11, color="2E75B6")
    r = line(r, "Prices and MAs are fetched by default. Use --offline to build from "
                "local JSON fallback data.", size=10, indent=2)
    r = line(r, "What IS refreshed live (Yahoo Finance / yfinance):",
             size=10, indent=2, color="375623")
    r = line(r, "    Prices  |  D200 SMA  |  W20/W50 weekly EMA  |  M20/M50 monthly EMA",
             size=10, indent=4, color="375623")
    r = line(r, "    TTM P/E  |  Forward P/E  |  D/E ratio  |  Net margin",
             size=10, indent=4, color="375623")
    r = line(r, "    Analyst consensus (Strong Buy/Buy/Hold/Sell)  |  Analyst mean price target",
             size=10, indent=4, color="375623")
    r = line(r, "Static portfolio data:", size=10, indent=2, color="7030A0")
    r = line(r, "    2022 bear lows  |  Cycle ATHs  |  Thesis / notes  |  Manual levels",
             size=10, indent=4, color="7030A0")
    r = line(r, "")
    r = line(r, "CLI OPTIONS", bold=True, size=11, color="2E75B6")
    r = line(r, "The script supports these optional command-line flags:", size=10, indent=2)
    r = line(r, "  --exclude TICKER [TICKER ...]   Omit one or more tickers from this run "
                "(no rows written)", size=10, indent=2, color="7030A0")
    r = line(r, "  --add    TICKER [TICKER ...]   Fetch a ticker live, append it, and "
                "persist it for future normal runs", size=10, indent=2, color="375623")
    r = line(r, "  --remove ASSET_OR_TICKER [...]   Persistently remove portfolio rows "
                "by exact asset name/ticker", size=10, indent=2, color="C00000")
    r = line(r, "  --offline   Build from local JSON fallback data without live market refresh",
             size=10, indent=2, color="7030A0")
    r = line(r, "  --output PATH   Write to a specific workbook path",
             size=10, indent=2, color="7030A0")
    r = line(r, "  --run-date ISO_DATETIME   Use a deterministic run date/time, "
                "e.g. 2026-07-28T09:30", size=10, indent=2, color="7030A0")
    r = line(r, "  --data-dir PATH   Read/write JSON sidecar data from a specific folder",
             size=10, indent=2, color="7030A0")
    r = line(r, "Examples:", size=10, indent=2)
    for example in (
        f"{COMMAND} --exclude NKE INTC",
        f"{COMMAND} --add ORCL ADBE",
        f"{COMMAND} --remove SpaceX",
        f"{COMMAND} --remove SPCX",
        f"{COMMAND} --add ORCL --exclude INTC",
        f"{COMMAND} --offline --output /tmp/workbook.xlsx --run-date 2026-07-28T09:30",
    ):
        r = line(r, f"  {example}", size=10, indent=4, color="1F4E79")
    r = line(r, "--add requires live refresh and internet access. Successful additions are "
                "saved in workbook/data/portfolio_rows.json and refreshed on "
                "future runs.", size=10, indent=2)
    r = line(r, "--remove deletes matching rows from workbook/data/portfolio_rows.json.",
             size=10, indent=2)
    r = line(r, "")
    r = line(r, "NOT FINANCIAL ADVICE — for personal tracking and research only.",
             bold=True, size=9, color="888888")

    return ws
