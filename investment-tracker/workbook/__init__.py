"""Workbook rendering, split by sheet.

Renderers consume prepared portfolio rows plus a ``SheetContext``. They never
fetch market data or touch JSON.
"""

from openpyxl import Workbook

from workbook import alert_levels, help_sheet, methodology, summary
from workbook.context import SheetContext

__all__ = ["SheetContext", "build_workbook"]

# Visible tab order. Daily Summary leads because it is the actionable view.
SHEET_ORDER = [
    summary.SHEET_TITLE,
    alert_levels.SHEET_TITLE,
    methodology.SHEET_TITLE,
    help_sheet.SHEET_TITLE,
]


def build_workbook(rows, macro_swing_rows, ctx):
    """Render every sheet and return the finished workbook.

    Alert Levels is built first because it owns the default sheet, then Daily
    Summary is inserted at index 0 so it becomes the first visible tab.
    """
    wb = Workbook()

    alert_levels.render(wb.active, rows, ctx)
    methodology.render(wb.create_sheet(methodology.SHEET_TITLE), macro_swing_rows)
    summary.render(wb.create_sheet(summary.SHEET_TITLE, 0), rows, ctx)
    help_sheet.render(wb.create_sheet(help_sheet.SHEET_TITLE))

    return wb
