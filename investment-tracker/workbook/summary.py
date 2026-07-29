"""Daily Summary sheet — the first tab.

Shows every priced asset grouped by Fib zone: BUY ZONES → APPROACHING →
WATCHING → NO FIB. Derived from the same rows as Alert Levels, then sorted by
actionability so the first tab shows what needs attention rather than mirroring
the full dashboard order.
"""

from openpyxl.utils import get_column_letter

from modules.calculations import (
    classify_zone,
    first_sentence,
    upside_to_ath,
)
from workbook.styles import (
    AL1_BG,
    AL2_BG,
    AL3_BG,
    AMB,
    D4_BG,
    DARK,
    GRN,
    RED_C,
    aln,
    bdr,
    fill,
    fnt,
)

SHEET_TITLE = "Daily Summary"

COLS = [
    "Asset", "Ticker", "CCY",
    "Current\nPrice",
    "Zone / Action",
    "AL1\n38.2%", "AL2\n50.0%", "AL3\n61.8%", "Deep Value\n78.6%",
    "Gap to\nAL1",
    "Upside\nto ATH",
    "Analyst\nConsensus",
    "Analyst\nPrice Target",
    "Key Thesis",
]
WIDTHS = [26, 8, 5, 11, 20, 12, 12, 12, 12, 10, 10, 14, 13, 52]

# Zone rank: 0=deepest, 1=AL3, 2=AL2, 3=AL1, 4=approaching, 5=above, 9=no fib
ZONE_STYLES = {
    # rank: (background, font colour, label)
    0: ("880000", "FFFFFF", "BELOW 78.6%  DEEPEST VALUE"),
    1: ("FFCDD2", "880000", "AL3 HIT  BACK UP TRUCK"),
    2: ("FFE0B2", "BF360C", "AL2 HIT  STRONG BUY"),
    3: ("E8F5E9", "1B5E20", "AL1 HIT  ACCUMULATE"),
    4: ("FFF9C4", "F57F17", "APPROACHING AL1"),
    5: ("F5F5F5", "757575", "Above AL1 — Watching"),
    9: ("ECEFF1", "607D8B", "No Fib / ETF"),
}

# Non-tradeable sentiment gauges show zone only — no action verb, because we
# cannot actually buy them. Colour coding is unchanged.
ZONE_LABELS_SENTIMENT = {
    0: "BELOW 78.6%",
    1: "AL3 HIT",
    2: "AL2 HIT",
    3: "AL1 HIT",
    4: "APPROACHING AL1",
    5: "Above AL1",
    9: "No Fib / ETF",
}

SECTION_DEFS = [
    (range(0, 4),
     "IN BUY ZONES  —  Price has retraced to a Fib alert level — action required today",
     "B71C1C"),
    (range(4, 5),
     "APPROACHING AL1  —  Within 15% of the 38.2% accumulate level — set price alerts",
     "E65100"),
    (range(5, 6),
     "WATCHING  —  Above AL1 — no entry yet; monitor for pullback to alert levels",
     "37474F"),
    (range(9, 10),
     "NO FIB / ETF  —  Manual levels or ETF; refer to Alert Levels tab",
     "455A64"),
]


def section_group_key(zone_rank):
    """Return the section index for a zone rank, so ranks 0-3 share one header."""
    for idx, (grp_range, _label, _bg) in enumerate(SECTION_DEFS):
        if zone_rank in grp_range:
            return idx
    return None


def build_summary_rows(rows, ctx):
    """Derive sorted summary records from portfolio rows.

    Section rows and rows without a current price are omitted because the
    summary is an actionable price/zone view.
    """
    summary = []
    for record in rows:
        if record.ticker in ctx.excluded_tickers:
            continue
        if record.is_section or record.current is None:
            continue

        zone = classify_zone(
            record.current, record.macro_lo, record.macro_hi,
            manual=record.manual, is_manual=record.is_manual,
        )
        summary.append({
            "row": record,
            "zone_rank": zone.rank, "al1": zone.al1, "al2": zone.al2,
            "al3": zone.al3, "d4": zone.d4, "gap": zone.gap_to_al1,
            "upside": upside_to_ath(record.current, record.macro_hi),
            "analyst": ctx.analyst_for(record.ticker),
            "note": first_sentence(record.notes),
        })

    # Within buy zones (0-3) the most extreme gap sorts first; approaching (4)
    # sorts closest first; the rest fall through in load order.
    summary.sort(key=lambda x: (
        x["zone_rank"],
        x["gap"] if x["gap"] is not None else 99,
    ))
    return summary


def _write_header(ws, ctx):
    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 95

    ws.merge_cells("A1:N1")
    ws["A1"] = "SIPP / ISA — Daily Action Summary  |  Fibonacci Alert Levels"
    ws["A1"].font = fnt(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = fill(DARK)
    ws["A1"].alignment = aln()
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:N2")
    ws["A2"] = (
        f"{ctx.run_date}  |  "
        "IN BUY ZONES: price has retraced to AL1/AL2/AL3 — action required  |  "
        "APPROACHING: within 15% of AL1 — set alerts  |  "
        "WATCHING: above AL1 — no action  |  "
        "Gap to AL1: negative = already in zone (price below AL1)  |  "
        "Upside = % from current to cycle ATH"
    )
    ws["A2"].font = fnt(italic=True, color="555555", size=7.5)
    ws["A2"].fill = fill("F0F0F0")
    ws["A2"].alignment = aln()
    ws.row_dimensions[2].height = 13

    for ci, (heading, width) in enumerate(zip(COLS, WIDTHS), 1):
        c = ws.cell(row=3, column=ci, value=heading)
        c.font = fnt(bold=True, color="FFFFFF", size=7.5)
        c.fill = fill(DARK)
        c.alignment = aln()
        c.border = bdr()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 30


def _write_section_row(ws, row, text, bg):
    ws.merge_cells(f"A{row}:N{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(bold=True, color="FFFFFF", size=8)
    c.fill = fill(bg)
    c.alignment = aln("left")
    c.border = bdr()
    ws.row_dimensions[row].height = 14


def _write_footer(ws, row, ctx):
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"] = (
        f"Prices: {ctx.data_source}  |  "
        "Fib: 2022 bear market low → cycle ATH  |  "
        "AL1=38.2%  AL2=50%  AL3=61.8%  Deep Value=78.6%  |  "
        "Gap to AL1: negative = price already below AL1 (in zone)  |  "
        "Not financial advice."
    )
    ws[f"A{row}"].font = fnt(italic=True, color="888888", size=7)
    ws[f"A{row}"].fill = fill("F0F0F0")
    ws[f"A{row}"].alignment = aln()
    ws.row_dimensions[row].height = 12


def render(ws, rows, ctx):
    """Write the Daily Summary sheet from prepared portfolio rows."""
    _write_header(ws, ctx)
    summary = build_summary_rows(rows, ctx)

    row = 4
    last_group = None

    for entry in summary:
        record = entry["row"]
        rank = entry["zone_rank"]
        group_key = section_group_key(rank)

        if group_key != last_group:
            if group_key is not None:
                _range, label, bg = SECTION_DEFS[group_key]
                _write_section_row(ws, row, label, bg)
                row += 1
            last_group = group_key

        zone_bg, zone_fc, zone_label = ZONE_STYLES[rank]
        in_buy_zone = rank <= 3
        row_bg = zone_bg if in_buy_zone else ("FFFFFF" if row % 2 == 0 else "F9F9F9")

        def cell(col, val, bold=False, color="000000", num_fmt=None, align="center"):
            c = ws.cell(row=row, column=col, value=val)
            c.font = fnt(bold=bold, color=color, size=8)
            c.fill = fill(row_bg)
            c.alignment = aln(align)
            c.border = bdr()
            if num_fmt:
                c.number_format = num_fmt
            return c

        cell(1, record.asset, bold=in_buy_zone,
             color=(zone_fc if in_buy_zone else "000000"), align="left")
        cell(2, record.ticker, bold=True,
             color=(zone_fc if in_buy_zone else "0000FF"))
        cell(3, record.ccy)
        cell(4, record.current, bold=True, color="0000FF", num_fmt='#,##0.00')

        # E — zone label, colour-coded
        display_label = (
            ZONE_LABELS_SENTIMENT[rank]
            if record.ticker in ctx.sentiment_tickers else zone_label
        )
        c = ws.cell(row=row, column=5, value=display_label)
        c.font = fnt(bold=in_buy_zone, color=zone_fc, size=7.5)
        c.fill = fill(zone_bg)
        c.alignment = aln()
        c.border = bdr()

        # F/G/H/I — alert levels
        for col, val, al_bg, al_fc in (
            (6, entry["al1"], AL1_BG, GRN),
            (7, entry["al2"], AL2_BG, AMB),
            (8, entry["al3"], AL3_BG, RED_C),
            (9, entry["d4"], D4_BG, "880000"),
        ):
            if val:
                c = ws.cell(row=row, column=col, value=val)
                c.font = fnt(size=8, bold=in_buy_zone, color=al_fc)
                c.number_format = '#,##0.00'
            else:
                c = ws.cell(row=row, column=col, value="—")
                c.font = fnt(size=8, color="888888")
            c.fill = fill(al_bg)
            c.alignment = aln()
            c.border = bdr()

        # J — gap to AL1
        gap = entry["gap"]
        if gap is not None:
            if gap < 0:
                gap_bg, gap_fc = AL1_BG, GRN      # in zone
            elif gap < 0.08:
                gap_bg, gap_fc = AL2_BG, AMB      # very close
            else:
                gap_bg, gap_fc = "F5F5F5", "757575"
            c = ws.cell(row=row, column=10, value=f"{gap:+.1%}")
            c.font = fnt(size=8, bold=(gap < 0), color=gap_fc)
            c.fill = fill(gap_bg)
        else:
            c = ws.cell(row=row, column=10, value="—")
            c.font = fnt(size=8, color="888888")
            c.fill = fill("F5F5F5")
        c.alignment = aln()
        c.border = bdr()

        # K — upside to ATH
        upside = entry["upside"]
        if upside is not None:
            if upside > 0.30:
                up_bg, up_fc = AL1_BG, GRN
            elif upside > 0.10:
                up_bg, up_fc = "FFF8E1", AMB
            else:
                up_bg, up_fc = "F5F5F5", "757575"
            c = ws.cell(row=row, column=11, value=f"{upside:+.1%}")
            c.font = fnt(size=8, color=up_fc)
            c.fill = fill(up_bg)
        else:
            c = ws.cell(row=row, column=11, value="—")
            c.font = fnt(size=8, color="888888")
            c.fill = fill("F5F5F5")
        c.alignment = aln()
        c.border = bdr()

        # L/M — analyst consensus and price target
        analyst = entry["analyst"]
        is_etf = record.is_etf
        if analyst and not is_etf:
            consensus = analyst.consensus
            if consensus == "Strong Buy":
                cons_bg, cons_fc = AL1_BG, GRN
            elif consensus == "Buy":
                cons_bg, cons_fc = "E8F5E9", "1B5E20"
            elif consensus == "Hold":
                cons_bg, cons_fc = AL2_BG, AMB
            else:
                cons_bg, cons_fc = "FFCDD2", RED_C
            c = ws.cell(row=row, column=12, value=f"{consensus} ({analyst.n_analysts})")
            c.font = fnt(size=8, bold=(consensus in ("Strong Buy", "Buy")), color=cons_fc)
            c.fill = fill(cons_bg)
        else:
            c = ws.cell(row=row, column=12, value=record.na_label)
            c.font = fnt(size=8, color="888888")
            c.fill = fill("F5F5F5")
        c.alignment = aln()
        c.border = bdr()

        if analyst and not is_etf and record.current:
            pt_bold = False
            pt_upside = analyst.upside_from(record.current)
            if pt_upside is not None:
                pt_label = f"${analyst.target:,.0f} ({pt_upside:+.0%})"
                if pt_upside > 0.30:
                    pt_bg, pt_fc, pt_bold = AL1_BG, GRN, True
                elif pt_upside > 0.10:
                    pt_bg, pt_fc = "E8F5E9", "1B5E20"
                elif pt_upside >= 0:
                    pt_bg, pt_fc = AL2_BG, AMB
                else:
                    pt_bg, pt_fc = "FFCDD2", RED_C
            else:
                pt_label, pt_bg, pt_fc = "N/A", "F5F5F5", "888888"
            c = ws.cell(row=row, column=13, value=pt_label)
            c.font = fnt(size=8, bold=pt_bold, color=pt_fc)
            c.fill = fill(pt_bg)
        else:
            c = ws.cell(row=row, column=13, value="N/A")
            c.font = fnt(size=8, color="888888")
            c.fill = fill("F5F5F5")
        c.alignment = aln()
        c.border = bdr()

        # N — key thesis (first clause of the note)
        c = ws.cell(row=row, column=14, value=entry["note"])
        c.font = fnt(size=7.5, italic=(rank >= 5))
        c.fill = fill(row_bg)
        c.alignment = aln("left")
        c.border = bdr()

        ws.row_dimensions[row].height = 28
        row += 1

    _write_footer(ws, row, ctx)
    return ws
