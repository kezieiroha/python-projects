"""Alert Levels sheet — the detailed dashboard.

Column map:
    A=Asset  B=Ticker  C=CCY  D=Current
    E=2022 Bear Low  F=Cycle ATH  G=Fib Range
    H=AL1/38.2%  I=AL2/50%  J=AL3/61.8%  K=78.6%
    L=Upside  M=Status
    N=TTM_PE  O=TTM_EPS  P=Fwd_PE  Q=Fwd_EPS  R=DE  S=Margin
    T=W20 EMA  U=W50 EMA  V=M20 EMA  W=M50 EMA  X=D200 SMA  Y=MA Trend
    Z=Thesis / Notes  AA=Analyst Consensus  AB=Analyst Price Target

All Fib levels are written as snapshot values. Historical anchor edits belong in
workbook/data/portfolio_rows.json, not in generated workbook cells.
"""

from openpyxl.utils import get_column_letter

from modules.calculations import (
    classify_zone,
    fib,
    ma_trend_text,
    upside_to_ath,
)
from modules.models import RowType
from workbook.styles import (
    AL1_BG,
    AL2_BG,
    AL3_BG,
    ALT,
    AMB,
    BLUE,
    D4_BG,
    DARK,
    FIB_BG,
    FUND_BG,
    GRN,
    ISA_BG,
    MA_BG,
    MA_HEAD,
    MANUAL,
    PURPLE,
    RED_C,
    SECT,
    SIPP_BG,
    WHITE,
    aln,
    bdr,
    fill,
    fnt,
    ma_cell_color,
    ma_num_fmt,
)

SHEET_TITLE = "Alert Levels"

COLS = [
    "Asset", "Ticker", "CCY",
    "Current\nPrice",
    "2022 Bear\nLow", "Cycle\nATH", "Fib\nRange",
    "AL1 — Accumulate\n38.2% Retrace",
    "AL2 — Strong Buy\n50.0% Retrace",
    "AL3 — Back Up Truck\n61.8% Retrace",
    "Deep Value\n78.6% Retrace",
    "Upside\nto ATH",
    "Status\nvs Levels",
    "TTM\nP/E", "TTM\nEPS", "Fwd\nP/E", "Fwd\nEPS", "D/E", "Net\nMargin",
    "W20\nEMA", "W50\nEMA", "M20\nEMA", "M50\nEMA",
    "D200\nSMA",
    "MA\nTrend",
    "Thesis / Notes",
    "Analyst\nConsensus",
    "Analyst\nPrice Target",
]
WIDTHS = [27, 9, 5, 11, 11, 11, 9, 14, 14, 15, 13, 9, 12, 8, 8, 8, 8, 7, 8,
          9, 9, 9, 9, 9, 12, 52, 13, 14]

GROUP_FILLS = {
    (1, 4): DARK,
    (5, 7): "37474F",
    (8, 11): PURPLE,
    (12, 13): "1B5E20",
    (14, 19): "00695C",   # fundamentals: TTM P/E, TTM EPS, Fwd P/E, Fwd EPS, D/E, Net Margin
    (20, 25): MA_HEAD,    # MAs: W20, W50, M20, M50, D200, MA Trend
    (26, 26): DARK,       # Thesis/Notes
    (27, 28): "283593",   # indigo — Analyst Consensus (AA) + Price Target (AB)
}

AL_CONFIG = [
    (8, 0.382, AL1_BG, GRN),
    (9, 0.500, AL2_BG, AMB),
    (10, 0.618, AL3_BG, RED_C),
    (11, 0.786, D4_BG, "880000"),
]


def _write_header(ws, ctx):
    ws.title = SHEET_TITLE
    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 90

    ws.merge_cells("A1:AB1")
    ws["A1"] = (
        "SIPP / ISA Investment Dashboard — Fibonacci Retracement Alert Levels  |  v21"
    )
    ws["A1"].font = fnt(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = fill(DARK)
    ws["A1"].alignment = aln()
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:AB2")
    ws["A2"] = (
        f"v21 | {ctx.run_date} | {ctx.data_source} | "
        "MACRO FIB: anchored to 2022 bear market low → cycle ATH  |  "
        "D200 SMA: simple 200-day MA; break below = regime change  |  "
        "AL1=38.2%  AL2=50.0%  AL3=61.8%  |  "
        "FINANCIALS: TTM P/E | TTM EPS | Fwd P/E | Fwd EPS | D/E | Net Margin  "
        "EPS: $x.xx = Alpha Vantage independent verification  "
        "~$x.xx italic = yfinance or derived fallback  "
        "Red = rule broken (D/E>1.0 / margin<0 / TTM P/E>100 / Fwd P/E>80 / negative EPS)  "
        "Amber = warning (D/E 0.75-1.0 / margin 0-5% / Fwd P/E 40-80)  Green = healthy  |  "
        "Teal cols = EMA/SMA: Green=above, Red=below  |  "
        "ANALYST: Consensus + mean price target from available live/fallback sources  "
        "Strong Buy=dark green  Buy=green  Hold=amber  Sell=red  |  "
        "PT: >30% upside=green  10-30%=light green  0-10%=amber  negative=red"
    )
    ws["A2"].font = fnt(italic=True, color="555555", size=7.5)
    ws["A2"].fill = fill("F0F0F0")
    ws["A2"].alignment = aln()
    ws.row_dimensions[2].height = 13

    for ci, heading in enumerate(COLS, 1):
        c = ws.cell(row=3, column=ci, value=heading)
        c.font = fnt(bold=True, color="FFFFFF", size=7.5)
        c.alignment = aln()
        c.border = bdr()
        for (lo, hi), colour in GROUP_FILLS.items():
            if lo <= ci <= hi:
                c.fill = fill(colour)
                break

    ws.row_dimensions[3].height = 40
    for ci, width in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width


def _row_background(row, sheet_row):
    if row.rtype == RowType.ISA:
        return ISA_BG
    if row.rtype == RowType.SIPP:
        return SIPP_BG
    if row.is_manual:
        return MANUAL
    return ALT if sheet_row % 2 == 0 else WHITE


def financial_flags(row, macro_tickers):
    """Red-flag rules. Macro gauges and pre-IPO placeholders are not assessed.

    D/E>1.0 = HIGH LEVERAGE; margin<0 = LOSS-MAKING; 0-5% = THIN MARGIN;
    TTM P/E>100 = ELEVATED; Fwd P/E>80 = STRETCHED.
    """
    if row.ticker in macro_tickers or row.rtype not in RowType.EQUITY_LIKE:
        return []

    flags = []
    if row.de is not None:
        if row.de > 2.0:
            flags.append(f"EXTREME LEVERAGE D/E={row.de:.2f}")
        elif row.de > 1.0:
            flags.append(f"HIGH LEVERAGE D/E={row.de:.2f}")
    if row.margin is not None:
        if row.margin < 0:
            flags.append("LOSS-MAKING")
        elif row.margin < 0.05:
            flags.append(f"THIN MARGIN {row.margin:.0%}")
    if row.ttm_pe is not None and row.ttm_pe > 100:
        flags.append(f"HIGH TTM P/E {row.ttm_pe:.0f}x")
    if row.fwd_pe is not None and row.fwd_pe > 80:
        flags.append(f"FWD P/E STRETCHED {row.fwd_pe:.0f}x")
    return flags


def _write_section_row(ws, row, asset):
    ws.merge_cells(f"A{row}:AB{row}")
    c = ws.cell(row=row, column=1, value=asset)
    c.font = fnt(bold=True, color="FFFFFF", size=8)
    c.fill = fill(SECT)
    c.alignment = aln("left")
    c.border = bdr()
    ws.row_dimensions[row].height = (
        63 if asset.startswith("MARKET SENTIMENT") else 15
    )


def _write_footer(ws, row, ctx):
    ws.merge_cells(f"A{row}:AB{row}")
    ws[f"A{row}"] = (
        f"Sources: Yahoo Finance (yfinance) — prices + MAs auto-refreshed "
        f"{ctx.run_date}  |  Analyst consensus: Alpha Vantage (manual refresh)  |  "
        "MACRO FIB: 2022 bear market monthly Low → cycle ATH monthly High  |  "
        "D200 SMA: simple 200-day rolling mean on daily close  |  "
        "W20/W50 EMA: daily resampled to weekly (Fri), EWM  |  "
        "M20/M50 EMA: 10Y monthly close, EWM  |  "
        "Green=price above MA (bullish), Red=price below MA (bearish)  |  "
        "GENIUS Act (Jul 2025): stablecoin yield prohibited — OCC rules Mar 2026  |  "
        "Not financial advice."
    )
    ws[f"A{row}"].font = fnt(italic=True, color="888888", size=7)
    ws[f"A{row}"].fill = fill("F0F0F0")
    ws[f"A{row}"].alignment = aln()
    ws.row_dimensions[row].height = 12


def _styled(ws, row, col, value, *, bg, color="000000", bold=False,
            italic=False, size=8, num_fmt=None, align="center"):
    """Write one bordered, filled, aligned cell. Every column writer uses this."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = fnt(bold=bold, color=color, size=size, italic=italic)
    c.fill = fill(bg)
    c.alignment = aln(align)
    c.border = bdr()
    if num_fmt:
        c.number_format = num_fmt
    return c


def _write_identity(ws, row, record, bg, fin_flags):
    """A-D: asset, ticker, currency, current price.

    A broken financial rule prefixes the asset with a bold red "! " so the row
    is scannable without reading the Status column.
    """
    if fin_flags:
        label, color, bold = f"! {record.asset}", RED_C, True
    elif record.rtype == RowType.ISA:
        label, color, bold = record.asset, BLUE, True
    elif record.rtype == RowType.SIPP:
        label, color, bold = record.asset, "000000", True
    else:
        label, color, bold = record.asset, "000000", False

    _styled(ws, row, 1, label, bg=bg, color=color, bold=bold, align="left")
    _styled(ws, row, 2, record.ticker, bg=bg, bold=True)
    _styled(ws, row, 3, record.ccy, bg=bg)
    _styled(ws, row, 4, record.current, bg=bg, bold=True, color="0000FF",
            num_fmt='#,##0.00' if record.current else None)


def _write_fib_columns(ws, row, record):
    """E-K: historical anchors, the range, and the four alert levels.

    Anchors are never touched by the live overlay. Manual rows carry explicit
    levels; everything else derives them from the anchors.
    """
    for col, anchor in ((5, record.macro_lo), (6, record.macro_hi)):
        _styled(ws, row, col, anchor, bg=FIB_BG, color="0000FF",
                num_fmt='#,##0.00' if anchor else None)

    if record.has_fib_anchors:
        _styled(ws, row, 7, round(record.macro_hi - record.macro_lo, 2),
                bg=FIB_BG, num_fmt='#,##0.00')
    else:
        _styled(ws, row, 7, "—", bg=FIB_BG)

    for col, pct, al_bg, al_color in AL_CONFIG:
        if record.is_manual and record.manual:
            value = record.manual[col - 8] if col <= 10 else "N/A"
            _styled(ws, row, col, value, bg=al_bg, color=al_color, bold=True,
                    italic=True,
                    num_fmt='#,##0.00' if isinstance(value, (int, float)) and value
                    else None)
        elif record.has_fib_anchors:
            _styled(ws, row, col, fib(record.macro_hi, record.macro_lo, pct),
                    bg=al_bg, color=al_color, num_fmt='#,##0.00')
        else:
            _styled(ws, row, col, "—", bg=al_bg)


def _write_upside(ws, row, record):
    """L: gain required to reach the cycle ATH."""
    if record.current and record.macro_hi:
        _styled(ws, row, 12, upside_to_ath(record.current, record.macro_hi),
                bg="E8F5E9", bold=True, num_fmt='0.0%')
    else:
        _styled(ws, row, 12, "—", bg="E8F5E9")


def _write_status(ws, row, record, bg, fin_flags):
    """M: zone status, with any financial red flags appended."""
    if record.is_manual:
        status, color = "Manual levels\n(no macro Fib)", AMB
    elif record.current and record.has_fib_anchors:
        zone = classify_zone(record.current, record.macro_lo, record.macro_hi)
        status, color = zone.status, zone.color
    else:
        status, color = "—", "555555"

    if fin_flags:
        status = status + "\n" + " | ".join(fin_flags)
        if color == "555555":    # was "No action yet" — upgrade to amber
            color = AMB

    _styled(ws, row, 13, status, bg=bg, color=color, size=7.5,
            bold=("HIT" in status or "BELOW" in status or bool(fin_flags)))


def _ratio_style(value, high, mid, low):
    """Shared banding for the P/E columns: red above high, amber above mid,
    green below low, neutral otherwise."""
    if value > high:
        return "FFCDD2", RED_C
    if value > mid:
        return AL2_BG, AMB
    if value < low:
        return AL1_BG, GRN
    return FUND_BG, "000000"


def _write_valuation(ws, row, record, ctx):
    """N-S: TTM P/E, TTM EPS, Fwd P/E, Fwd EPS, D/E, net margin."""
    na = record.na_label
    current, ttm_pe, fwd_pe = record.current, record.ttm_pe, record.fwd_pe

    # N — TTM P/E
    if ttm_pe is not None:
        pe_bg, pe_fc = _ratio_style(ttm_pe, 100, 50, 15)
        _styled(ws, row, 14, f"{ttm_pe:.0f}x", bg=pe_bg, color=pe_fc,
                bold=ttm_pe > 100)
    else:
        _styled(ws, row, 14, na, bg=FUND_BG, color="888888")

    # O — TTM EPS. Source priority:
    #   1. Alpha Vantage OVERVIEW (DilutedEPSTTM) — independent verification
    #   2. yfinance trailingEps — live fallback, not independent verification
    #   3. Derived Price ÷ TTM P/E — "~" prefix, cannot verify the P/E
    eps_view = ctx.eps_for(record.ticker)
    ttm_eps = eps_view.trailing_eps if eps_view else None
    verified = bool(eps_view and eps_view.is_verified)
    if ttm_eps is None and ttm_pe and current:
        ttm_eps, verified = round(current / ttm_pe, 2), False

    if ttm_eps is not None:
        prefix = "$" if verified else "~$"
        _styled(
            ws, row, 15,
            f"{'-' if ttm_eps < 0 else ''}{prefix}{abs(ttm_eps):.2f}",
            bg="FFCDD2" if ttm_eps < 0 else FUND_BG,
            color=RED_C if ttm_eps < 0 else ("000000" if verified else "555555"),
            italic=not verified,
        )
    else:
        _styled(ws, row, 15, na, bg=FUND_BG, color="888888")

    # P — Fwd P/E
    if fwd_pe is not None:
        fpe_bg, fpe_fc = _ratio_style(fwd_pe, 80, 40, 20)
        _styled(ws, row, 16, f"{fwd_pe:.0f}x", bg=fpe_bg, color=fpe_fc,
                bold=fwd_pe > 80)
    else:
        _styled(ws, row, 16, na, bg=FUND_BG, color="888888")

    # Q — Fwd EPS is always derived; there is no independent forward source.
    fwd_eps = round(current / fwd_pe, 2) if (fwd_pe and current) else None
    if fwd_eps is not None:
        _styled(ws, row, 17, f"{'-' if fwd_eps < 0 else ''}~${abs(fwd_eps):.2f}",
                bg="FFCDD2" if fwd_eps < 0 else FUND_BG,
                color=RED_C if fwd_eps < 0 else "555555", italic=True)
    else:
        _styled(ws, row, 17, na, bg=FUND_BG, color="888888")

    # R — D/E (red >1.0, amber 0.75-1.0, green <0.50)
    de = record.de
    if de is not None:
        if de > 1.0:
            de_bg, de_fc = "FFCDD2", RED_C
        elif de > 0.75:
            de_bg, de_fc = AL2_BG, AMB
        elif de < 0.50:
            de_bg, de_fc = AL1_BG, GRN
        else:
            de_bg, de_fc = FUND_BG, "000000"
        _styled(ws, row, 18, f"{de:.2f}", bg=de_bg, color=de_fc, bold=de > 1.0)
    else:
        _styled(ws, row, 18, na, bg=FUND_BG, color="888888")

    # S — Net Margin (red <0, amber 0-5%, green >20%)
    margin = record.margin
    if margin is not None:
        if margin < 0:
            m_bg, m_fc, m_bold = "FFCDD2", RED_C, True
        elif margin < 0.05:
            m_bg, m_fc, m_bold = AL2_BG, AMB, False
        elif margin > 0.25:
            m_bg, m_fc, m_bold = AL1_BG, GRN, True
        elif margin > 0.10:
            m_bg, m_fc, m_bold = "E8F5E9", GRN, False
        else:
            m_bg, m_fc, m_bold = FUND_BG, "000000", False
        _styled(ws, row, 19, margin, bg=m_bg, color=m_fc, bold=m_bold,
                num_fmt='0%')
    else:
        _styled(ws, row, 19, na, bg=FUND_BG, color="888888")


def _write_moving_averages(ws, row, record, ctx):
    """T-Y: the four EMAs, the D200 SMA, and a compact trend summary."""
    mas = ctx.moving_averages_for(record.ticker)
    for col, ema in ((20, mas.w20), (21, mas.w50), (22, mas.m20),
                     (23, mas.m50), (24, mas.d200)):
        if ema is not None:
            cell_bg, cell_fc = ma_cell_color(record.current, ema)
            _styled(ws, row, col, ema, bg=cell_bg, color=cell_fc,
                    num_fmt=ma_num_fmt(ema))
        else:
            _styled(ws, row, col, "N/A", bg=MA_BG, color="888888")

    _styled(
        ws, row, 25,
        ma_trend_text(record.current, mas.w20, mas.w50, mas.m20, mas.m50, mas.d200),
        bg=MA_BG, size=7.5,
    )


def _write_thesis_and_analyst(ws, row, record, ctx, bg):
    """Z-AB: thesis note, analyst consensus, and analyst price target."""
    _styled(ws, row, 26, record.notes, bg=bg, size=7.5,
            italic=(record.rtype == RowType.ISA), align="left")

    analyst = ctx.analyst_for(record.ticker)
    if not (analyst and not record.is_etf):
        # ETF, crypto without analyst data, UK stocks, pre-IPO
        for col in (27, 28):
            _styled(ws, row, col, record.na_label, bg="F5F5F5", color="888888")
        return

    consensus = analyst.consensus
    if consensus == "Strong Buy":
        cons_bg, cons_fc = AL1_BG, GRN
    elif consensus == "Buy":
        cons_bg, cons_fc = "E8F5E9", "1B5E20"
    elif consensus == "Hold":
        cons_bg, cons_fc = AL2_BG, AMB
    else:
        cons_bg, cons_fc = "FFCDD2", RED_C
    _styled(ws, row, 27, f"{consensus} ({analyst.n_analysts})",
            bg=cons_bg, color=cons_fc,
            bold=(consensus in ("Strong Buy", "Buy")))

    upside = analyst.upside_from(record.current)
    if upside is None:
        _styled(ws, row, 28, "N/A", bg="F5F5F5", color="888888")
        return

    if upside > 0.30:
        pt_bg, pt_fc, pt_bold = AL1_BG, GRN, True
    elif upside > 0.10:
        pt_bg, pt_fc, pt_bold = "E8F5E9", "1B5E20", False
    elif upside >= 0:
        pt_bg, pt_fc, pt_bold = AL2_BG, AMB, False
    else:
        pt_bg, pt_fc, pt_bold = "FFCDD2", RED_C, False
    _styled(ws, row, 28, f"${analyst.target:,.0f} ({upside:+.0%})",
            bg=pt_bg, color=pt_fc, bold=pt_bold)


def render(ws, rows, ctx):
    """Write the Alert Levels sheet from prepared portfolio rows."""
    _write_header(ws, ctx)

    row = 4
    for record in rows:
        if record.ticker in ctx.excluded_tickers:
            continue

        if record.is_section:
            _write_section_row(ws, row, record.asset)
            row += 1
            continue

        bg = _row_background(record, row)
        fin_flags = financial_flags(record, ctx.macro_tickers)

        _write_identity(ws, row, record, bg, fin_flags)
        _write_fib_columns(ws, row, record)
        _write_upside(ws, row, record)
        _write_status(ws, row, record, bg, fin_flags)
        _write_valuation(ws, row, record, ctx)
        _write_moving_averages(ws, row, record, ctx)
        _write_thesis_and_analyst(ws, row, record, ctx, bg)

        ws.row_dimensions[row].height = 32
        row += 1

    _write_footer(ws, row, ctx)
    return ws
