"""Fib Methodology sheet.

Explains the alert-level framework and shows curated historical macro swing
examples from workbook/data/macro_swing_data.json. These rows are illustrative; the
operational alert rows live on Alert Levels and Daily Summary.
"""

from openpyxl.utils import get_column_letter

from modules.calculations import fib
from workbook.styles import (
    AL1_BG,
    AL2_BG,
    AL3_BG,
    AMB,
    D4_BG,
    DARK,
    GRN,
    RED_C,
    WHITE,
    aln,
    bdr,
    fill,
    fnt,
)

SHEET_TITLE = "Fib Methodology"
WIDTHS = [22, 14, 14, 14, 13, 55]

# Worked example uses META's macro swing: 2022 low 87.40 -> cycle ATH 794.38.
EXAMPLE_LOW = 87.40
EXAMPLE_HIGH = 794.38

LEVELS = [
    ("AL1 — Accumulate", "38.2%", "ATH−(Range×0.382)", 0.382,
     "First macro pullback",
     "Start building. Institutional first stop. ~1/3 position.",
     AL1_BG, GRN),
    ("AL2 — Strong Buy", "50.0%", "ATH−(Range×0.500)", 0.500,
     "Macro midpoint",
     "Increase size. 50% retrace = major support. ~1/3 more.",
     AL2_BG, AMB),
    ("AL3 — Back Up Truck", "61.8%", "ATH−(Range×0.618)", 0.618,
     "Golden ratio retrace",
     "Maximum conviction. Final third. Historically rare at macro scale.",
     AL3_BG, RED_C),
    ("Deep Value", "78.6%", "ATH−(Range×0.786)", 0.786,
     "Near full retrace",
     "Extreme caution OR maximum size if thesis intact. Structural breakdown risk.",
     D4_BG, "880000"),
]


def render(ws, macro_swing_rows):
    """Write the Fib Methodology sheet."""
    def write(row, values, bold=False, bg=None, colors=None):
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=val)
            colour = colors[ci - 1] if colors and ci <= len(colors) else "000000"
            c.font = fnt(bold=bold, color=colour, size=8.5)
            c.fill = fill(bg or WHITE)
            c.alignment = aln("left")
            c.border = bdr()
        ws.row_dimensions[row].height = 18

    for ci, width in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.merge_cells("A1:F1")
    ws["A1"] = "Macro Fibonacci Retracement Framework — Anchor Methodology"
    ws["A1"].font = fnt(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = fill(DARK)
    ws["A1"].alignment = aln()
    ws.row_dimensions[1].height = 22

    row = 2
    write(row, ["Level", "Fib %", "Formula", "Example (META)",
                "Trigger Meaning", "Interpretation"],
          bold=True, bg="37474F", colors=["FFFFFF"] * 6)
    row += 1

    for label, pct, formula, ratio, trigger, interpretation, bg, fc in LEVELS:
        example = f"${fib(EXAMPLE_HIGH, EXAMPLE_LOW, ratio):,.2f}"
        write(row, [label, pct, formula, example, trigger, interpretation],
              bg=bg, colors=[fc, fc, "000000", "000000", "000000", "000000"])
        row += 1

    row += 1
    write(row, ["MACRO SWING ANCHORS (from workbook/data/macro_swing_data.json)",
                "", "", "", "", ""], bold=True, bg="F5F5F5")
    row += 1
    write(row, ["Ticker", "2022 Bear Low", "Cycle ATH", "Current",
                "Macro AL1", "Macro Status"],
          bold=True, bg="37474F", colors=["FFFFFF"] * 6)
    row += 1

    for swing in macro_swing_rows:
        fc = swing.font_color
        write(
            row,
            [
                swing.ticker,
                f"{swing.macro_lo:,.2f}" if swing.macro_lo else "N/A",
                f"{swing.macro_hi:,.2f}" if swing.macro_hi else "N/A",
                f"{swing.current:,.2f}" if swing.current else "N/A",
                swing.al1_display,
                swing.status,
            ],
            bg=swing.bg,
            colors=["000000", "0000FF", "0000FF", "000000", fc, fc],
        )
        row += 1

    return ws
