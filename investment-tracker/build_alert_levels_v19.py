import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

_ts     = datetime.now().strftime("%d%m%y-%H%M")
_here   = os.path.dirname(os.path.abspath(__file__))
OUTPUT  = os.path.join(_here, f"SIPP_Alert_Levels_v21_{_ts}.xlsx")

# ── CLI arguments ─────────────────────────────────────────────────────────────
import argparse as _ap
_parser = _ap.ArgumentParser(description="SIPP/ISA Dashboard generator")
_parser.add_argument(
    "--exclude", nargs="+", metavar="TICKER", default=[],
    help="Tickers to omit from this run, e.g. --exclude NKE INTC"
)
_parser.add_argument(
    "--add", nargs="+", metavar="TICKER", default=[],
    help="Extra tickers to fetch live and append to Watchlist, e.g. --add ORCL ADBE"
)
_args = _parser.parse_args()
EXCLUDED_TICKERS = {t.upper() for t in _args.exclude}
ADD_TICKERS      = [t.upper() for t in _args.add]

wb = Workbook()

# ── Colours ───────────────────────────────────────────────────────────────────
DARK    = "1A1A2E"
SECT    = "16213E"
AL1_BG  = "E8F5E9"
AL2_BG  = "FFF8E1"
AL3_BG  = "FFEBEE"
D4_BG   = "FFCCCC"
FIB_BG  = "EDE7F6"
FUND_BG = "E0F2F1"
ISA_BG  = "E3F2FD"
SIPP_BG = "F3E5F5"
MANUAL  = "FFF3E0"
ALT     = "F5F5F5"
WHITE   = "FFFFFF"
GRN     = "2E7D32"
AMB     = "E65100"
RED_C   = "C62828"
BLUE    = "0D47A1"
PURPLE  = "4A148C"
MA_BG   = "E0F7FA"
MA_HEAD = "00695C"

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=8, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def aln(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────────
# Live Market Data Refresh
# ─────────────────────────────────────────────────────────────────────────────
# When LIVE_REFRESH = True, current prices and all five MAs (W20/W50/M20/M50/D200)
# are fetched from Yahoo Finance via yfinance and merged over the hardcoded dicts.
# Hardcoded values remain as fallback if a fetch fails or yfinance is unavailable.
# One-time install:  pip3 install yfinance pandas
# ─────────────────────────────────────────────────────────────────────────────
LIVE_REFRESH = True

# ── Alpha Vantage — independent EPS source (verifies yfinance P/E ratios) ────
# Set the AV_KEY environment variable before running:
#   export AV_KEY="your_key_here"
# Free tier: 25 calls/day, 5/min → set AV_SLEEP=13 (default)
# Premium tier (75+/min) → set AV_SLEEP=1 or lower
AV_KEY   = os.environ.get("AV_KEY", "")
AV_SLEEP = 13    # seconds between COMPANY_OVERVIEW calls

# US equities to fetch from Alpha Vantage (ETFs, crypto, commodities excluded)
_AV_US = [
    "MSTR", "COIN", "CEG",  "XOM",  "COP",  "MRVL", "PLTR", "CCJ",
    "PANW", "RKLB", "NVDA", "AMZN", "MSFT", "META", "GOOGL", "NFLX",
    "SMR",  "LUNR", "TLN",  "GEV",  "ORCL", "AAPL", "TSLA", "INTC", "NKE",
    # Asymmetric Growth Basket
    "SDGR", "PENG", "PATH", "S",    "FLYW", "AMSC", "PRCT",
    "PGY",  "NTLA", "BEAM", "ABSI", "BKSY", "RCAT", "PL",
    "DNN",  "QS",   "IONQ", "QBTS", "RGTI",
]
# LSE equities: map script key -> Alpha Vantage .LON symbol
_AV_UK = {
    "NATO.L": "NATO.LON", "RR.L":   "RR.LON",   "NUCG.L": "NUCG.LON",
    "SHEL.L": "SHEL.LON", "BP.L":   "BP.LON",   "MNTN.L": "MNTN.LON",
}

live_prices       = {}   # script_ticker -> float
live_ma           = {}   # script_ticker -> (w20, w50, m20, m50, d200)
live_fundamentals = {}   # script_ticker -> (ttm_pe, fwd_pe, de, margin)
live_eps          = {}   # script_ticker -> (ttm_eps, is_derived)
live_analyst      = {}   # script_ticker -> (consensus_label, n_analysts, mean_pt)

# Script ticker keys that differ from Yahoo Finance symbols
_YAHOO = {"BTC": "BTC-USD", "VIX": "^VIX", "TNX": "^TNX", "DXY": "DX-Y.NYB",
          "GRA": "GRA.TO"}   # NanoXplore — TSX Venture Exchange

# All tickers with obtainable live prices (TBC / CRCL excluded — no live price yet)
_FETCH = [
    "MSTR",  "COIN",  "CEG",   "XOM",   "COP",   "MRVL",  "PLTR",  "CCJ",
    "PANW",  "RKLB",  "NVDA",  "AMZN",  "MSFT",  "META",  "GOOGL", "NFLX",
    "SMR",   "LUNR",  "TLN",   "GEV",   "ORCL",  "AAPL",  "TSLA",  "INTC",
    "NKE",   "XRP-USD", "ETH-USD", "BTC", "BMNR",
    "NATO.L","RR.L",  "NUCG.L","SHEL.L","BP.L",
    "FGRD.L","VPNG.L","IUSU.L","VWRP.L","SMGB.L","VUSA.L","SMT.L", "MNTN.L",
    "CL=F",  "VIX",  "TNX",  "TLT",  "DXY",
    # Asymmetric Growth Basket
    "SDGR", "PENG", "PATH", "S",    "FLYW", "AMSC", "PRCT",
    "PGY",  "NTLA", "BEAM", "ABSI", "BKSY", "RCAT", "PL",
    "DNN",  "GRA",  "QS",   "IONQ", "QBTS", "RGTI",
]

if LIVE_REFRESH:
    try:
        import yfinance as _yf
        import pandas as _pd

        def _yt(t):
            """Map script ticker key to Yahoo Finance symbol."""
            return _YAHOO.get(t, t)

        def _rnd(v):
            """Round to 4dp for sub-$1 assets, 2dp otherwise. Returns None on NaN."""
            if v is None:
                return None
            try:
                f = float(v)
                return None if f != f else round(f, 4 if abs(f) < 1 else 2)
            except Exception:
                return None

        _syms = [_yt(t) for t in _FETCH]
        import io as _io, contextlib as _cl

        # ── Daily 2Y — MA calculations only (D200/W20/W50) ──────────────────
        print(f"[Live Refresh] Fetching daily data ({len(_syms)} tickers)...")
        with _cl.redirect_stdout(_io.StringIO()):
            _raw_d = _yf.download(_syms, period="2y",  interval="1d",
                                  auto_adjust=True, progress=False, group_by="ticker")

        # ── Hourly 5D — current prices (last completed hourly bar; avoids
        #    stale daily-close and .info caching issues) ──────────────────────
        print(f"[Live Refresh] Fetching hourly prices ({len(_syms)} tickers)...")
        with _cl.redirect_stdout(_io.StringIO()):
            _raw_h = _yf.download(_syms, period="5d",  interval="60m",
                                  auto_adjust=True, progress=False, group_by="ticker")

        # ── Monthly 10Y — M20/M50 calculations ──────────────────────────────
        print(f"[Live Refresh] Fetching monthly data ({len(_syms)} tickers)...")
        with _cl.redirect_stdout(_io.StringIO()):
            _raw_m = _yf.download(_syms, period="10y", interval="1mo",
                                  auto_adjust=True, progress=False, group_by="ticker")

        _multi = len(_syms) > 1

        for _t in _FETCH:
            _ys = _yt(_t)
            try:
                _dc = (_raw_d[_ys]["Close"] if _multi else _raw_d["Close"]).dropna()
                _hc = (_raw_h[_ys]["Close"] if _multi else _raw_h["Close"]).dropna()
                _mc = (_raw_m[_ys]["Close"] if _multi else _raw_m["Close"]).dropna()
                if _dc.empty:
                    print(f"  [skip] {_t}: no daily data returned")
                    continue

                # Current price — hourly bar is most up-to-date batch source;
                # falls back to daily close if hourly data unavailable
                _cur_src = _hc if not _hc.empty else _dc
                live_prices[_t] = _rnd(_cur_src.iloc[-1])

                # D200 SMA (need at least 50 bars before it becomes meaningful)
                _d200 = _rnd(_dc.rolling(200).mean().iloc[-1]) if len(_dc) >= 50 else None

                # Weekly EMAs — resample daily to weekly Friday close
                _wk  = _dc.resample("W-FRI").last().dropna()
                _w20 = _rnd(_wk.ewm(span=20, adjust=False).mean().iloc[-1]) if len(_wk) >= 10 else None
                _w50 = _rnd(_wk.ewm(span=50, adjust=False).mean().iloc[-1]) if len(_wk) >= 20 else None

                # Monthly EMAs
                _m20 = _rnd(_mc.ewm(span=20, adjust=False).mean().iloc[-1]) if len(_mc) >= 10 else None
                _m50 = _rnd(_mc.ewm(span=50, adjust=False).mean().iloc[-1]) if len(_mc) >= 20 else None

                live_ma[_t] = (_w20, _w50, _m20, _m50, _d200)

            except Exception as _e:
                print(f"  [warn] {_t}: {_e}")

        # ── Fundamentals + Analyst data (per-ticker .info calls) ─────────────
        _rec_map = {
            "strong_buy":     "Strong Buy",
            "buy":            "Buy",
            "outperform":     "Buy",
            "overweight":     "Buy",
            "hold":           "Hold",
            "neutral":        "Hold",
            "market_perform": "Hold",
            "equal_weight":   "Hold",
            "underperform":   "Sell",
            "sell":           "Sell",
            "underweight":    "Sell",
            "strong_sell":    "Strong Sell",
        }
        print(f"[Live Refresh] Fetching fundamentals + analyst data ({len(_FETCH)} tickers)...")
        for _t in _FETCH:
            _ys = _yt(_t)
            try:
                _info = _yf.Ticker(_ys).info

                def _to_float(v):
                    """Safely coerce yfinance value to float.
                    Returns None on failure, on NaN, and on ±Inf."""
                    try:
                        f = float(v)
                        return None if (f != f or abs(f) == float("inf")) else f
                    except (TypeError, ValueError):
                        return None

                # TTM P/E — None for loss-makers (negative) or unavailable
                _ttm = _to_float(_info.get("trailingPE"))
                if _ttm is not None and _ttm <= 0:
                    _ttm = None
                _ttm = round(_ttm, 1) if _ttm is not None else None

                # Fwd P/E — None for negative estimates or unavailable
                _fpe = _to_float(_info.get("forwardPE"))
                if _fpe is not None and _fpe <= 0:
                    _fpe = None
                _fpe = round(_fpe, 1) if _fpe is not None else None

                # D/E — primary: debtToEquity (returned as percentage, e.g. 97.9 = 0.979)
                # Fallback: compute from totalDebt / (bookValue * sharesOutstanding)
                _de_raw = _to_float(_info.get("debtToEquity"))
                if _de_raw is None:
                    _t_debt  = _to_float(_info.get("totalDebt"))
                    _bv      = _to_float(_info.get("bookValue"))
                    _shares  = _to_float(_info.get("sharesOutstanding")
                                         or _info.get("impliedSharesOutstanding"))
                    if _t_debt is not None and _bv and _shares and (_bv * _shares) > 0:
                        # Express as percentage to match debtToEquity convention
                        _de_raw = (_t_debt / (_bv * _shares)) * 100
                _de = round(_de_raw / 100, 2) if _de_raw is not None else None

                # Net margin — primary: profitMargins (decimal, e.g. 0.0484)
                # Fallback: netIncomeToCommon / totalRevenue
                _mg = _to_float(_info.get("profitMargins"))
                if _mg is None:
                    _ni  = _to_float(_info.get("netIncomeToCommon"))
                    _rev = _to_float(_info.get("totalRevenue"))
                    if _ni is not None and _rev and _rev > 0:
                        _mg = _ni / _rev
                _mg = round(_mg, 4) if _mg is not None else None

                live_fundamentals[_t] = (_ttm, _fpe, _de, _mg)

                # EPS — trailingEps (TTM) and forwardEps (next 12m analyst estimate)
                _ttm_eps = _to_float(_info.get("trailingEps"))
                _fwd_eps = _to_float(_info.get("forwardEps"))
                if _ttm_eps is not None:
                    _ttm_eps = round(_ttm_eps, 2)
                if _fwd_eps is not None:
                    _fwd_eps = round(_fwd_eps, 2)
                live_eps[_t] = (_ttm_eps, _fwd_eps)

                # Current price — use regularMarketPrice from .info (real-time)
                # Overrides the daily-close bar used for MA calc which can be 1+ days stale
                _rt_price = (_info.get("regularMarketPrice")
                             or _info.get("currentPrice")
                             or _info.get("navPrice"))
                if _rt_price is not None:
                    live_prices[_t] = _rnd(float(_rt_price))

                # Analyst consensus
                _rec_key = (_info.get("recommendationKey") or "").lower()
                _n_ana   = _info.get("numberOfAnalystOpinions")
                _mean_pt = _info.get("targetMeanPrice")
                _cons    = _rec_map.get(_rec_key)
                if _cons and _n_ana and _mean_pt:
                    live_analyst[_t] = (_cons, int(_n_ana), round(float(_mean_pt), 2))

            except Exception as _e:
                print(f"  [warn] {_t} fundamentals: {_e}")

        print(f"[Live Refresh] Done — {len(live_prices)} prices | "
              f"{len(live_ma)} MA sets | "
              f"{len(live_fundamentals)} fundamentals | "
              f"{len(live_analyst)} analyst sets.")

    except ImportError:
        print("[Live Refresh] yfinance not installed.")
        print("[Live Refresh] Run:  pip3 install yfinance pandas  then re-run the script.")
    except Exception as _ex:
        print(f"[Live Refresh] Failed ({_ex}) — falling back to hardcoded values.")

# ── Alpha Vantage EPS fetch ───────────────────────────────────────────────────
# Independent TTM EPS from COMPANY_OVERVIEW — used to verify yfinance P/E ratios.
# Only runs when AV_KEY is set and LIVE_REFRESH is True.
# Skips tickers where AV has no data (ETFs, crypto, commodities, pre-IPO).
# live_eps entry: (ttm_eps, is_derived=False) for AV data.
if AV_KEY and LIVE_REFRESH:
    try:
        import requests as _req
        import time    as _time

        _av_targets = list(_AV_US) + list(_AV_UK.keys())
        print(f"[Alpha Vantage EPS] Fetching {len(_av_targets)} tickers "
              f"(sleep={AV_SLEEP}s/call)...")

        for _t in _av_targets:
            _av_sym = _AV_UK.get(_t, _t)   # remap .L → .LON for UK stocks
            try:
                _resp = _req.get(
                    "https://www.alphavantage.co/query",
                    params={"function": "OVERVIEW", "symbol": _av_sym, "apikey": AV_KEY},
                    timeout=15,
                )
                _d = _resp.json()

                # Detect rate-limit or empty response
                if "Note" in _d or "Information" in _d or "Symbol" not in _d:
                    _msg = _d.get("Note") or _d.get("Information") or "no data"
                    print(f"  [AV skip] {_t}: {_msg[:80]}")
                    _time.sleep(AV_SLEEP)
                    continue

                # TTM EPS — DilutedEPSTTM is the authoritative field
                _eps_raw = _d.get("DilutedEPSTTM") or _d.get("EPS")
                try:
                    _ttm_eps = round(float(_eps_raw), 2) if _eps_raw else None
                except (TypeError, ValueError):
                    _ttm_eps = None

                if _ttm_eps is not None:
                    live_eps[_t] = (_ttm_eps, False)   # False = real AV data, not derived
                    print(f"  [AV] {_t}: TTM EPS=${_ttm_eps}")

            except Exception as _e:
                print(f"  [AV warn] {_t}: {_e}")

            _time.sleep(AV_SLEEP)

        print(f"[Alpha Vantage EPS] Done — {len(live_eps)} EPS values fetched.")

    except ImportError:
        print("[Alpha Vantage EPS] requests library not installed. "
              "Run:  pip3 install requests")
    except Exception as _ex:
        print(f"[Alpha Vantage EPS] Failed ({_ex})")

# ── MA / EMA / D200 data ───────────────────────────────────────────────────────
# All from Yahoo Finance (daily 1Y + monthly 10Y/5Y).  31 Mar 2026.
# Tuple: (W20 EMA, W50 EMA, M20 EMA, M50 EMA, D200 SMA)
# W20/W50 = daily resampled to weekly (Fri close), EWM.
# M20/M50 = 10Y monthly close, EWM.
# D200    = 200-day rolling simple mean on daily close.
# ── Analyst consensus + price targets (Alpha Vantage COMPANY_OVERVIEW, 10 Apr 2026) ──
# Consensus score: SB=1, B=2, H=3, S=4, SS=5; weighted mean
# 1.0-1.5=Strong Buy  1.5-2.5=Buy  2.5-3.5=Hold  3.5+=Sell
# Format: ticker -> (consensus_label, num_analysts, mean_price_target)
analyst_data = {
    "MSTR":  ("Buy",  14,  374.07),
    "COIN":  ("Buy",  33,  240.22),
    "CEG":   ("Buy",  20,  375.82),
    "XOM":   ("Hold", 26,  160.17),
    "COP":   ("Buy",  29,  130.86),
    "MRVL":  ("Buy",  43,  121.25),
    "PLTR":  ("Hold", 28,  185.25),
    "CCJ":   ("Buy",  22,  128.68),
    "PANW":  ("Buy",  57,  205.96),
    "RKLB":  ("Buy",  16,   87.58),
    "NVDA":  ("Buy",  63,  268.22),
    "AMZN":  ("Buy",  68,  281.27),
    "MSFT":  ("Buy",  58,  587.31),
    "META":  ("Buy",  67,  860.25),
    "GOOGL": ("Buy",  67,  376.29),
    "NFLX":  ("Buy",  51,  113.43),
    "SMR":   ("Hold", 16,   20.73),
    "LUNR":  ("Buy",   9,   22.50),
    "TLN":   ("Buy",  16,  467.84),
    "GEV":   ("Buy",  34,  887.89),
    "ORCL":  ("Buy",  38,  192.50),   # Alpha Vantage limit hit — consensus estimate Apr 2026
    "AAPL":  ("Buy",  52,  255.00),   # Alpha Vantage limit hit — consensus estimate Apr 2026
    "TSLA":  ("Hold", 57,  285.00),   # consensus estimate Apr 2026
    "INTC":  ("Hold", 50,   48.00),   # consensus estimate Apr 2026 (pre-surge; PT lags price)
    "NKE":   ("Buy",  39,   63.64),   # Alpha Vantage Apr 2026; SB=5 B=19 H=13 S=1 SS=1
}

ma_data = {
    "MSTR":    (168.99,    230.27,    219.43,    167.66,    259.53),
    "COIN":    (217.36,    239.74,    234.14,    210.39,    281.92),
    "CEG":     (313.36,    304.46,    283.91,    207.07,    328.82),
    "XOM":     (139.96,    126.05,    120.45,    102.75,    121.29),
    "COP":     (107.91,     99.98,    100.56,     92.76,     96.40),
    "MRVL":    ( 84.47,     78.80,     79.67,     68.57,     80.93),
    "PLTR":    (155.56,    148.86,    122.82,     77.52,    164.02),
    "CCJ":     (104.04,     87.97,     78.51,     55.61,     91.94),
    "PANW":    (172.64,    180.54,    174.16,    146.29,    188.50),
    "RKLB":    ( 66.86,     54.79,     42.93,     26.40,     57.23),
    "NVDA":    (179.82,    168.75,    151.72,    105.44,     91.67),
    "AMZN":    (217.70,    217.23,    209.56,    180.29,    167.90),
    "MSFT":    (429.31,    447.79,    432.41,    374.55,    373.19),
    "META":    (638.49,    645.37,    609.51,    485.12,    436.05),
    "GOOGL":   (297.37,    261.30,    234.26,    180.12,    163.87),
    "NFLX":    ( 94.70,    100.58,     94.93,     75.31,    107.65),
    "BMNR":    ( 26.31,       None,      None,      None,     36.63),
    "XRP-USD": (  1.70,       2.03,      1.80,      1.37,      2.06),
    "ETH-USD": (2510.83,   2734.96,  2798.53,  2502.58,   3092.88),
    "BTC":     (79353.94,  88689.36, 83485.94, 65485.30,  91314.70),
    "NATO.L":  (  19.54,     18.46,     16.47,      None,     19.01),
    "RR.L":    (1193.27,   1091.67,    924.26,    612.43,   1130.46),
    "NUCG.L":  (  44.36,     39.44,     35.63,      None,     41.77),
    "SHEL.L":  (2982.58,   2833.11,   2770.18,   2487.07,   2786.04),
    "BP.L":    ( 480.95,    449.39,    447.67,    427.86,    440.45),
    "CL=F":    (  81.49,     71.70,     70.82,     70.98,     67.02),
    "VIX":     (  21.37,     20.92,      None,      None,     18.18),
    "FGRD.L":  (4160.65,   3867.77,   3682.18,      None,   3937.04),
    "VPNG.L":  (  15.72,     14.30,     13.48,     11.95,     14.34),
    "IUSU.L":  ( 823.92,    796.76,    765.07,    693.48,    794.52),
    "VWRP.L":  ( 126.06,    121.11,    116.29,    103.87,    122.60),
    "SMGB.L":  (  49.15,     43.07,     39.80,     31.69,     43.66),
    "VUSA.L":  (  94.95,     92.41,     89.19,     77.96,     94.06),   # 14 Apr 2026; D200 estimated from 200-day weighted avg
    "SMT.L":   (1175.02,   1115.69,   1087.50,   1025.30,   1130.22),
    "MNTN.L":  (   1.58,      1.36,      1.28,      1.25,      1.35),
    "TLN":     ( 351.91,    333.11,    291.03,      None,    363.09),   # W50,M50 N/A — Jul 2023 inception
    "GEV":     ( 737.88,    622.55,    541.82,      None,    649.82),   # M50 N/A — Apr 2024 spin-off
    "ORCL":    ( 169.63,    184.14,    173.18,    142.38,    216.27),
    "AAPL":    ( 257.26,    244.91,    239.38,    202.76,    250.31),
    "TSLA":    ( 393.10,    380.53,    351.53,    336.30,    397.35),   # 13 Apr 2026; M50 25mo approx
    "INTC":    (  46.06,     36.82,     36.03,     31.91,     36.41),   # 14 Apr 2026; D200 exact via Polygon SQL
    "NKE":     (  56.16,     59.90,     66.92,     80.31,     65.17),   # 15 Apr 2026; Yahoo Finance MCP
    "TNX":     (   4.42,      4.38,      4.35,      4.28,      4.37),   # 10Y yield; % units (W20/W50/M20/M50/D200)
    "TLT":     (  86.41,     89.54,     90.12,     94.88,     89.62),   # 15 Apr 2026; iShares 20Y+ Treasury
    "DXY":     ( 101.50,    104.20,    105.10,    106.80,    104.50),   # 15 Apr 2026; US Dollar Index (DX-Y.NYB)
}

# EPS fallback — populated entirely by live refresh; no hardcoded values needed
# Format: ticker -> (ttm_eps, fwd_eps)  — both in per-share currency units
eps_data = {}

# Merge live MA, analyst, and EPS data over hardcoded fallbacks
if live_ma:
    ma_data.update(live_ma)
if live_analyst:
    analyst_data.update(live_analyst)
if live_eps:
    eps_data.update(live_eps)


def ma_num_fmt(val):
    if val >= 10000: return '#,##0'
    elif val >= 1:   return '#,##0.00'
    else:            return '0.0000'


def ma_trend_text(current, w20, w50, m20, m50, d200):
    if current is None:
        return "N/A"
    lines = []
    if d200 is not None:
        lines.append(f"D200: {'P>SMA' if current > d200 else 'P<SMA'}")
    if w20 is not None and w50 is not None:
        p_w20 = "P>" if current > w20 else "P<"
        align = "20>50" if w20 > w50 else "20<50"
        lines.append(f"W: {p_w20}20 {align}")
    elif w20 is not None:
        lines.append(f"W: {'P>20' if current > w20 else 'P<20'} (no W50)")
    if m20 is not None and m50 is not None:
        p_m20 = "P>" if current > m20 else "P<"
        align = "20>50" if m20 > m50 else "20<50"
        lines.append(f"M: {p_m20}20 {align}")
    elif m20 is not None:
        lines.append(f"M: {'P>20' if current > m20 else 'P<20'} (no M50)")
    return "\n".join(lines) if lines else "N/A"


def ma_cell_color(current, ema_val):
    if current is None or ema_val is None:
        return MA_BG, "888888"
    if current > ema_val:
        return "C8E6C9", GRN
    else:
        return "FFCDD2", RED_C


def fib(high, low, pct):
    return round(high - ((high - low) * pct), 2)


# ─────────────────────────────────────────────────────────────────────────────
# Run-time date / data source label (used in banner and footer)
# ─────────────────────────────────────────────────────────────────────────────
_dt       = datetime.now()
_run_date = f"{_dt.day} {_dt.strftime('%b %Y')}"
_data_src = (f"Live via yfinance — {_dt.strftime('%d %b %Y %H:%M')}"
             if live_prices else "Hardcoded fallback data")

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Dashboard  (24 cols: A–X)
# A=Asset  B=Ticker  C=CCY  D=Current
# E=2022 Bear Low  F=Cycle ATH  G=Fib Range
# H=AL1/38.2%  I=AL2/50%  J=AL3/61.8%  K=78.6%
# L=Upside  M=Status
# N=TTM_PE  O=TTM_EPS  P=Fwd_PE  Q=Fwd_EPS  R=DE  S=Margin
# T=W20 EMA  U=W50 EMA  V=M20 EMA  W=M50 EMA  X=D200 SMA  Y=MA Trend
# Z=Thesis / Notes  AA=Analyst Consensus  AB=Analyst Price Target
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Alert Levels"
ws.freeze_panes = "A4"
ws.sheet_view.zoomScale = 90

COLS = [
    "Asset", "Ticker", "CCY",
    "Current\nPrice",
    "2022 Bear\nLow", "Cycle\nATH", "Fib\nRange",
    "AL1 — Accumulate\n38.2% Retrace",
    "AL2 — Strong Buy\n50.0% Retrace",
    "AL3 — Back Up Truck\n61.8% Retrace",
    "Deep Value\n78.6% Retrace",
    "Upside\nto ATH",
    "Status\nvs Levels",
    "TTM\nP/E", "TTM\nEPS", "Fwd\nP/E", "Fwd\nEPS", "D/E", "Net\nMargin",
    "W20\nEMA", "W50\nEMA", "M20\nEMA", "M50\nEMA",
    "D200\nSMA",
    "MA\nTrend",
    "Thesis / Notes",
    "Analyst\nConsensus",
    "Analyst\nPrice Target",
]
WIDTHS = [27, 9, 5, 11, 11, 11, 9, 14, 14, 15, 13, 9, 12, 8, 8, 8, 8, 7, 8,
          9, 9, 9, 9, 9, 12, 52, 13, 14]

ws.merge_cells("A1:AB1")
ws["A1"] = "SIPP / ISA Investment Dashboard — Fibonacci Retracement Alert Levels  |  v21"
ws["A1"].font = fnt(bold=True, color="FFFFFF", size=13)
ws["A1"].fill = fill(DARK)
ws["A1"].alignment = aln()
ws.row_dimensions[1].height = 22

ws.merge_cells("A2:AB2")
ws["A2"] = (
    f"v21 | {_run_date} | {_data_src} | MACRO FIB: anchored to 2022 bear market low → cycle ATH  |  "
    "D200 SMA: simple 200-day MA; break below = regime change  |  "
    "AL1=38.2%  AL2=50.0%  AL3=61.8%  |  "
    "FINANCIALS: TTM P/E | TTM EPS | Fwd P/E | Fwd EPS | D/E | Net Margin  "
    "EPS: $x.xx = Alpha Vantage (independent, verifies P/E)  ~$x.xx italic = derived Price÷P/E  "
    "Red = rule broken (D/E>1.0 / margin<0 / TTM P/E>100 / Fwd P/E>80 / negative EPS)  "
    "Amber = warning (D/E 0.75-1.0 / margin 0-5% / Fwd P/E 40-80)  Green = healthy  |  "
    "Teal cols = EMA/SMA: Green=above, Red=below  |  "
    "ANALYST: Consensus + mean price target (Alpha Vantage / estimates, 14 Apr 2026)  "
    "Strong Buy=dark green  Buy=green  Hold=amber  Sell=red  |  "
    "PT: >30% upside=green  10-30%=light green  0-10%=amber  negative=red"
)
ws["A2"].font = fnt(italic=True, color="555555", size=7.5)
ws["A2"].fill = fill("F0F0F0")
ws["A2"].alignment = aln()
ws.row_dimensions[2].height = 13

grp_fills = {
    (1,  4):  DARK,
    (5,  7):  "37474F",
    (8,  11): PURPLE,
    (12, 13): "1B5E20",
    (14, 19): "00695C",   # fundamentals: TTM P/E, TTM EPS, Fwd P/E, Fwd EPS, D/E, Net Margin
    (20, 25): MA_HEAD,    # MAs: W20, W50, M20, M50, D200, MA Trend
    (26, 26): DARK,       # Thesis/Notes
    (27, 28): "283593",   # indigo — Analyst Consensus (AA) + Price Target (AB)
}
for ci, h in enumerate(COLS, 1):
    c = ws.cell(row=3, column=ci, value=h)
    c.font = fnt(bold=True, color="FFFFFF", size=7.5)
    c.alignment = aln()
    c.border = bdr()
    for (lo, hi), fc in grp_fills.items():
        if lo <= ci <= hi:
            c.fill = fill(fc)
            break

ws.row_dimensions[3].height = 40
for ci, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# ─────────────────────────────────────────────────────────────────────────────
# rows_data: (asset, ticker, ccy, current, macro_lo, macro_hi,
#             ttm_pe, fwd_pe, de, margin, notes, rtype, manual_levels)
# macro_lo = 2022 bear market low (monthly Low col)
# macro_hi = cycle ATH (monthly High col)
# manual_levels = (al1, al2, al3) used for rows with rtype == MAN
# ─────────────────────────────────────────────────────────────────────────────
S = "SECTION"; MAN = "MANUAL"; STK = "STOCK"
ETF_T = "ETF"; ISA_T = "ISA"; SIPP_T = "SIPP"

rows_data = [
    # ── MARKET SENTIMENT ─────────────────────────────────────────────────────
    ("MARKET SENTIMENT — WTI Crude (CL=F) + CBOE VIX + US 10Y Yield (^TNX) + iShares 20Y+ Treasury (TLT) + US Dollar Index (DXY) as macro regime gauges.\n"
     "VIX < 15 = complacency | 15-20 = normal | 20-30 = elevated | > 30 = fear | > 40 = panic.\n"
     "TNX: < 3.5% = loose | 3.5-4.0% = neutral | 4.0-4.5% = tight | > 4.5% = restrictive.\n"
     "DXY: < 95 = weak USD (risk-on) | 95-100 = neutral | 100-105 = strong (headwind) | > 105 = very strong (global tightening).  Rising DXY = headwind for commodities / gold / EM / crypto / US multinational earnings.\n"
     "TLT moves INVERSE to yields — falling TLT = rising rates = headwind for equities / growth / crypto.\n"
     "CL=F Fib anchor: Dec 2025 low $54.98 \u2192 Apr 2026 ATH $119.48.",
     "", "", None, None, None, None, None, None, None, "", S, None),

    ("WTI Crude Oil (Front Month)", "CL=F", "USD", 97.05,
     54.98, 119.48,
     None, None, None, None,
     "ABOVE AL1 — WATCH. Macro Fib (Dec 2025 low $54.98 → Apr 2026 ATH $119.48): "
     "AL1=$94.85 | AL2=$87.23 | AL3=$79.61 | 78.6%=$68.78. "
     "Current $97.05 above AL1 ($94.85). Crude softened from spike high — near AL1 support zone. "
     "Geopolitical risk premium. Crashed $28 in a single day (Apr 8: $112 → $91) on demand destruction fears "
     "then partially recovered. Monitor: crude declining toward $95 = AL1 — risk-off signal for equities. "
     "D200=$67.02; price well above D200. W20=$81.49 / W50=$71.70 — strong bull trend.",
     STK, None),

    ("CBOE Volatility Index", "VIX", "USD", 18.60,
     None, None,
     None, None, None, None,
     "NORMAL-LOW ZONE. VIX at 18.60 — back inside normal range (15-20), falling from recent spike. "
     "VIX thresholds: < 15 = complacency (market top risk) | 15-20 = normal | "
     "20-30 = elevated fear (prepare buy lists) | 30+ = fear (buy equities) | 40+ = panic (maximum opportunity). "
     "1Y range: 13.38 (complacency) to 35.75 (fear peak). Current 21.19 = moderate fear. "
     "D200=18.18; VIX back near D200 — risk-off regime easing. W20=21.37 / W50=20.92. "
     "MANUAL LEVELS: VIX 20 = alert | VIX 30 = buy equities aggressively | VIX 40 = back up truck.",
     MAN, (20.00, 30.00, 40.00)),

    ("US 10-Year Treasury Yield", "TNX", "PCT", 4.35,
     None, None,
     None, None, None, None,
     "TIGHT ZONE (4.0-4.5%). TNX at ~4.35% — within restrictive range. "
     "Yield thresholds: < 3.5% = loose / risk-on | 3.5-4.0% = neutral | 4.0-4.5% = tight | > 4.5% = restrictive (risk-off). "
     "Rising yields = headwind for growth stocks, crypto, long-duration bonds. "
     "Fed pivot expectations: market pricing ~2 cuts in 2026. Watch for breakout above 4.5% (risk-off trigger) "
     "or drop below 4.0% (risk-on signal). "
     "MANUAL ALERT LEVELS: 3.50 = loose / pivot confirmed | 4.00 = neutral boundary | 4.50 = restrictive / risk-off.",
     MAN, (3.50, 4.00, 4.50)),

    ("iShares 20+ Year Treasury Bond ETF", "TLT", "USD", 85.00,
     88.02, 179.70,
     None, None, None, None,
     "BELOW 78.6% — DEEPEST VALUE / HIGHEST YIELD. TLT below its 2022 bear low ($88.02). "
     "Macro Fib (2022 bear low $88.02 → 2020 ATH $179.70): "
     "AL1=$144.76 | AL2=$133.86 | AL3=$122.96 | 78.6%=$108.82. "
     "Current ~$85 BELOW 78.6% — bond market pricing structurally higher rates. "
     "TLT moves INVERSE to 10Y yield: rising yield = falling TLT. "
     "ETF as macro gauge only — TLT recovery signals rate normalisation / Fed pivot. "
     "D200=$89.62; price well below D200 — sustained bear regime for long-duration bonds. "
     "W20=$86.41 / W50=$89.54 — price below all key MAs.",
     ETF_T, None),

    ("US Dollar Index", "DXY", "IDX", 99.80,
     89.21, 114.78,
     None, None, None, None,
     "AL3 ZONE — WEAK USD / DOLLAR IN DECLINE. Macro Fib (Jan 2021 low 89.21 → Sep 2022 ATH 114.78): "
     "AL1=104.99 | AL2=101.99 | AL3=98.99 | 78.6%=95.68. "
     "Current ~99.80 — in the AL3 zone (between 50% and 61.8% retracement). "
     "Weak dollar is TAILWIND for: commodities, gold, crypto, EM equities, US multinationals' overseas earnings. "
     "Falling DXY often correlates with risk-on rotation. "
     "DXY thresholds: < 95 = weak / risk-on | 95-100 = neutral-weak | 100-105 = strong (headwind) | > 105 = very strong (global tightening). "
     "D200=104.50; price well below D200 — confirmed dollar bear trend. "
     "W20=101.50 / W50=104.20 — price below all key MAs, bearish dollar alignment.",
     STK, None),

    # ── CRYPTO ──────────────────────────────────────────────────────────────
    ("CRYPTO SLEEVE  —  All Fib levels now anchored to 2022 bear market low → cycle ATH (macro swing).  "
     "BTC D200 SMA ($91,315) = regime line; current $68,779 confirmed below — bear regime active.  "
     "GENIUS ACT (Jul 2025): stablecoins legitimised. Stablecoin yield PROHIBITED (OCC Mar 2026).",
     "", "", None, None, None, None, None, None, None, "", S, None),

    ("Strategy (MicroStrategy)", "MSTR", "USD", 132.36,
     13.26, 543.00,
     None, None, 1.74, None,
     "BELOW 78.6% MACRO FIB — DEEPEST VALUE ZONE. $125.75 below 78.6% ($126.62). "
     "Macro Fib (2022 low $13.26 → ATH $543.00): AL1=$340.63 | AL2=$278.13 | AL3=$215.63 | 78.6%=$126.62. "
     "Current sits BELOW 78.6% — historically extreme. BTC proxy. "
     "GTC LIMIT ORDER PENDING: BUY 100 @ $114 (near absolute floor). "
     "LEVERAGE NOTE: D/E=1.74 — ~$8.2B convertible debt used to buy BTC. Risk: debt service if BTC collapses. "
     "D200 SMA=$259.53; price well below D200 — bear regime confirmed. "
     "W20=$168.99 / W50=$230.27 / M20=$219.43 / M50=$167.66 — price below all MAs.",
     ISA_T, None),

    ("Circle (Stablecoin)", "CRCL", "USD", None, None, None,
     None, None, None, None,
     "GENIUS ACT PRIMARY BENEFICIARY — USDC LEGITIMISED. Post-IPO — set alerts once price history establishes. "
     "USDC daily stablecoin volume $1T → $4T. Yield on stablecoins prohibited (Act). "
     "Market share + institutional adoption is primary thesis. Monitor $20-25 range.",
     MAN, (None, None, None)),

    ("Coinbase", "COIN", "USD", 172.62,
     31.83, 444.65,
     39.0, 30.0, 0.53, 0.18,
     "BELOW AL3 MACRO FIB — BACK UP TRUCK. $173.38 between AL3 ($189.53) and 78.6% ($120.17). "
     "Macro Fib (2022 low $31.83 → ATH $444.65): AL1=$286.95 | AL2=$238.24 | AL3=$189.53 | 78.6%=$120.17. "
     "GENIUS Act: exchange business broadly POSITIVE — regulatory clarity drives institutional volume. "
     "USDC yield-sharing secondary risk (~$200-300M of ~$7B). "
     "D200=$281.92; price well below D200. W20=$217.36 / W50=$239.74 — bearish alignment.",
     STK, None),

    # ── DEFENCE ──────────────────────────────────────────────────────────────
    ("DEFENCE", "", "", None, None, None, None, None, None, None, "", S, None),

    ("HANetf Future of Defence", "NATO.L", "GBP", 19.97,
     11.62, 21.58,
     None, None, None, None,
     "AL1 ZONE. Macro Fib (inception low £11.62 → ATH £21.58): AL1=£17.78 | AL2=£16.60 | AL3=£15.42. "
     "Current £19.36 — above AL1. NATO rearmament cycle (5% GDP by 2035). "
     "D200=£19.01; price just above D200 — holding the line. W20=£19.54 / W50=£18.46.",
     ETF_T, None),

    ("Rolls-Royce Holdings", "RR.L", "GBp", 1286.0,
     64.43, 1420.0,
     17.0, 28.0, 0.60, 0.28,  # D/E=0.60 gross; net debt turning positive (£6.25B cash vs £4.36B debt)
     "ABOVE AL1 — ABOVE ENTRY ZONE. Macro Fib (2022 low 64p → ATH 1420p): "
     "AL1=954p | AL2=742p | AL3=531p. "
     "Current 1121.5p well above AL1 (954p) — wait for deeper pullback. "
     "UK DEFENCE + SMR: Aero engines (Typhoon/F-35), naval propulsion, SMR programme. "
     "D200=1130p; current 1121.5p just below D200 — minor bearish signal. "
     "W20=1193p / W50=1092p. Price between W20 and W50 — watch closely.",
     STK, None),

    # ── NUCLEAR ──────────────────────────────────────────────────────────────
    ("NUCLEAR / URANIUM", "", "", None, None, None, None, None, None, None, "", S, None),

    ("VanEck Uranium & Nuclear", "NUCG.L", "GBP", 44.71,
     32.00, 53.61,
     None, None, None, None,
     "AL2 HIT — STRONG BUY ZONE. Macro Fib (inception low £32.00 → ATH £53.61): "
     "AL1=£45.35 | AL2=£42.81 | AL3=£40.27 | 78.6%=£36.62. "
     "Current £42.20 — BELOW AL2 (£42.81), above AL3. Strong buy zone. "
     "D200=£41.77; current just above D200. W20=£44.36 / W50=£39.44 — price between W20 and W50.",
     ETF_T, None),

    ("Constellation Energy", "CEG", "USD", 291.72,
     40.73, 411.68,
     41.0, 22.0, 0.56, 0.09,
     "ABOVE AL1 — ABOVE ENTRY ZONE. Macro Fib (2022 low $40.73 → ATH $411.68): "
     "AL1=$270.53 | AL2=$226.21 | AL3=$181.88 | 78.6%=$119.57. "
     "Current $295.19 above AL1 ($270.53) — not at entry yet. Wait for AL1 test. "
     "Largest US nuclear fleet (24,000MW). Three Mile Island restarted. "
     "D200=$328.82; price below D200 — bearish signal. W20=$313.36 / W50=$304.46.",
     STK, None),

    ("NuScale Power", "SMR", "USD", 9.58,
     10.15, 57.42,
     None, None, 0.00, -4.50,  # Pre-revenue; deep operating losses (~450% net loss on minimal revenue)
     "WATCHLIST ONLY — HIGHLY SPECULATIVE. $10.30 near inception low. "
     "All macro Fib levels far above current (78.6%=$20.37). "
     "DO NOT BUY until firm project contracts. Monitor only.",
     STK, None),

    ("Talen Energy", "TLN", "USD", 326.08,
     49.50, 451.28,
     None, 10.66, 6.25, -0.0834,  # No TTM PE (loss-making); Fwd P/E=10.66; D/E=6.25 EXTREME LEVERAGE; margin=-8.34%
     "ABOVE AL1 — WATCH (HIGH RISK). Macro Fib (inception Jul 2023 low $49.50 -> ATH $451.28): "
     "AL1=$297.80 | AL2=$250.39 | AL3=$202.98 | 78.6%=$135.48. "
     "Current $324.54 above AL1 ($297.80). Amazon nuclear supply deal — Susquehanna plant (960MW, 20yr PPA). "
     "Emerged from bankruptcy 2023; capital structure carries extreme leverage (D/E=6.25). Net margin negative. "
     "D200=$363.09; price below D200 — bearish. W20=$351.91 / W50=$333.11 — price below both weekly EMAs. "
     "High risk/reward nuclear power pure-play. Set accumulate alert at AL1 ($297.80).",
     STK, None),

    ("GE Vernova", "GEV", "USD", 991.12,
     118.56, 948.38,
     48.0, 37.0, 0.11, 0.1283,  # TTM PE=48x; Fwd P/E=37x; D/E=0.11 net cash positive; margin=12.83%
     "ABOVE AL1 — WAIT FOR PULLBACK. Macro Fib (inception Apr 2024 low $118.56 -> ATH $948.38): "
     "AL1=$631.39 | AL2=$533.47 | AL3=$435.55 | 78.6%=$296.14. "
     "Current $853.16 well above AL1 ($631.39) — near ATH zone, not a buy. "
     "GE power/grid spin-off. Revenue $34B+. Gas turbines, onshore/offshore wind, grid infrastructure, SMR pipeline. "
     "Net cash positive (D/E=0.11). Net margin 12.83%. Fwd P/E 37x. "
     "D200=$649.82; price well above D200 — strong uptrend. W20=$737.88 / W50=$622.55 — price above both. "
     "Set alert at AL1 ($631.39) for a major pullback entry (~26% below current).",
     STK, None),

    # ── CRUDE OIL / ENERGY ────────────────────────────────────────────────────
    ("CRUDE OIL / ENERGY — INTEGRATED MAJORS + COMMODITY ETP", "", "", None, None, None, None, None, None, None, "", S, None),

    ("Shell PLC", "SHEL.L", "GBp", 3443.5,
     1619.12, 3490.0,
     15.5, 11.3, 0.40, 0.07,
     "NEAR ATH — WAIT FOR PULLBACK. 3435.5p — only 54p from ATH. "
     "Macro Fib (2022 low 1619p → ATH 3490p): AL1=2775p | AL2=2555p | AL3=2334p. "
     "Current well above all Fib levels — at ATH, not a buy zone. "
     "Set alert at AL1 (2775p) for ~20% pullback. "
     "D200=2786p; current above D200. W20=2983p / W50=2833p.",
     STK, None),

    ("BP PLC", "BP.L", "GBp", 575.2,
     329.80, 584.10,
     8.0, 13.0, 0.86, 0.00,  # TTM PE=8x (oil major, low multiple); D/E=0.86 elevated; margin 0% (write-downs)
     "NEAR ATH — WAIT FOR PULLBACK. 575.6p — only 8.5p from ATH. "
     "Macro Fib (2022 low 329.8p → ATH 584.1p): AL1=487p | AL2=457p | AL3=427p. "
     "Current above all Fib levels — near ATH. Set alert at AL1 (487p). "
     "D200=440p; price well above D200 — bullish. Fwd P/E 13x. Div 4.34%. CEO reset to fossil fuels.",
     STK, None),

    ("ExxonMobil", "XOM", "USD", 152.64,
     52.75, 171.23,
     24.0, 18.3, 0.19, 0.09,
     "NEAR ATH — WAIT FOR PULLBACK. $165.43 — only $5.80 from ATH. "
     "Macro Fib (2022 low $52.75 → ATH $171.23): AL1=$117.00 | AL2=$112.00 | AL3=$98.40 | 78.6%=$78.04. "
     "Current well above all Fib levels — near ATH. Set alert at AL1 ($117). "
     "D200=$121.29; price above D200. W20=$139.96 / W50=$126.05.",
     STK, None),

    ("ConocoPhillips", "COP", "USD", 123.62,
     61.71, 134.87,
     20.0, 17.5, 0.38, 0.13,
     "AT ATH — WAIT. $133.25 — only $1.62 from ATH ($134.87). "
     "Macro Fib (2022 low $61.71 → ATH $134.87): AL1=$106.92 | AL2=$98.29 | AL3=$79.60. "
     "Current well above all Fib levels. Macro not useful at ATH — wait for $107 area. "
     "D200=$96.40; price well above D200. Pure E&P, highest crude leverage.",
     STK, None),

    # ── AI INFRASTRUCTURE ─────────────────────────────────────────────────────
    ("AI INFRASTRUCTURE STACK", "", "", None, None, None, None, None, None, None, "", S, None),

    ("FT Smart Grid Infrastructure", "FGRD.L", "GBP", 45.87,
     21.22, 45.88,
     None, None, None, None,
     "ABOVE AL1. Macro Fib (2022 low £21.22 → ATH £45.88): AL1=£36.46 | AL2=£33.55 | AL3=£30.64. "
     "Current £42.30 well above AL1. Set alert at AL1 (£36.46). Layer 2: Power equipment.",
     ETF_T, None),

    ("Global X Data Centre REITs", "VPNG.L", "GBP", 17.86,
     8.67, 18.54,
     None, None, None, None,
     "ABOVE AL1. Macro Fib (2022 low £8.67 → ATH £18.54): AL1=£14.59 | AL2=£13.61 | AL3=£12.62. "
     "Current £16.07 above AL1. Set alert at AL1 (£14.59). "
     "D200=£14.34; price above D200. Layer 3: Data centre REITs.",
     ETF_T, None),

    ("iShares S&P 500 Utilities", "IUSU.L", "GBX", 848.75,
     530.0, 883.25,
     None, None, None, None,
     "NEAR ATH. Macro Fib (2022 low 530p → ATH 883p): AL1=748p | AL2=707p | AL3=665p. "
     "Current 839.5p above AL1 (748p) — above entry zone. "
     "D200=795p; price above D200. W20=824p / W50=797p. Layer 4: Utilities (Constellation, NextEra).",
     ETF_T, None),

    ("Marvell Technology", "MRVL", "USD", 131.30,
     34.65, 126.99,
     32.0, 18.0, 0.32, 0.33,
     "ABOVE AL1. Macro Fib (2022 low $34.65 → ATH $126.99): AL1=$91.71 | AL2=$80.82 | AL3=$70.92. "
     "Current $97.68 above AL1 ($91.71). Set accumulate alert at AL1 ($91.71). "
     "Custom AI silicon (XPUs) for AWS/Google/MSFT. D200=$80.93. W20=$84.47 / W50=$78.80.",
     STK, None),

    # ── WATCHLIST ─────────────────────────────────────────────────────────────
    ("WATCHLIST — INDIVIDUAL STOCKS", "", "", None, None, None, None, None, None, None, "", S, None),

    ("Palantir", "PLTR", "USD", 132.37,
     5.92, 207.52,
     242.0, 83.0, 0.00, 0.36,
     "ABOVE AL1 — WATCH. Macro Fib (2022 low $5.92 → ATH $207.52): "
     "AL1=$130.51 | AL2=$106.72 | AL3=$82.92 | 78.6%=$48.78. "
     "Current $133.58 — just above AL1 ($130.51). AT ACCUMULATE ZONE. Set alert at AL1. "
     "No debt. 36% net margin. Rev +70% YoY. D200=$164.02; current below D200 — bearish. "
     "W20=$155.56 / W50=$148.86 — price below both weekly EMAs.",
     STK, None),

    ("Cameco", "CCJ", "USD", 116.70,
     17.86, 135.24,
     111.0, 59.0, 0.15, 0.17,
     "ABOVE AL1. Macro Fib (2022 low $17.86 → ATH $135.24): AL1=$90.42 | AL2=$76.55 | AL3=$62.66. "
     "Current $104.67 above AL1 ($90.42). Set alert at AL1 ($90.42). "
     "Western uranium monopoly. D200=$91.94; current above D200. W20=$104.04 / W50=$87.97.",
     STK, None),

    ("Nike, Inc.", "NKE", "USD", 44.20,
     77.22, 166.63,
     29.0, 22.0, 0.98, 0.0484,
     "BELOW 78.6% MACRO FIB — DEEPEST VALUE. $44.20 below 78.6% ($96.35). "
     "Macro Fib (Oct 2022 low $77.22 → ATH $166.63 Nov 2021): AL1=$132.47 | AL2=$121.92 | AL3=$111.37 | 78.6%=$96.35. "
     "EXTREME: current price is BELOW the 2022 bear market low ($77.22) — historic dislocation (-43% below). "
     "CEO Elliott Hill (Oct 2024, replaced Donahoe) executing turnaround: DTC rebalancing, wholesale channel rebuild, brand reinvestment. "
     "Revenue +0.1% YoY; earnings -34.8% YoY. China slowdown structural headwind. "
     "Manufacturing risk: ~100% Asia-based production exposed to US tariffs. "
     "Net margin 4.8% (thin). D/E=0.98 (approaching high leverage). Div yield 3.7%. "
     "D200=$65.17; price massively below D200 — deep bear regime. "
     "W20=$56.16 / W50=$59.90 / M20=$66.92 / M50=$80.31 — price below ALL MAs. "
     "Only accumulate on confirmed recovery above 2022 bear low ($77.22). Set alert at $77.22 reclaim.",
     STK, None),

    ("CYBERSECURITY — AI-NATIVE PLATFORMS", "", "", None, None, None, None, None, None, None, "", S, None),

    ("Palo Alto Networks", "PANW", "USD", 162.51,
     68.37, 223.61,
     86.0, 39.0, 0.04, 0.13,
     "AL1 HIT — ACCUMULATE. Macro Fib (2022 low $68.37 → ATH $223.61): "
     "AL1=$164.31 | AL2=$146.00 | AL3=$127.68 | 78.6%=$101.49. "
     "Current $156.36 BELOW AL1 ($164.31), ABOVE AL2 ($146.00) — ACCUMULATE ZONE. "
     "AI-native cybersecurity (Cortex XDR, XSIAM, SASE). Net margin 13%. Fwd P/E 39x. "
     "D200=$188.50; price well below D200. W20=$172.64 / W50=$180.54 — bearish.",
     STK, None),

    ("SPACE / LUNAR ECONOMY", "", "", None, None, None, None, None, None, None, "", S, None),

    ("Rocket Lab", "RKLB", "USD", 70.62,
     3.48, 99.58,
     None, None, 0.00, -0.26,  # Pre-profit; net margin approx -26% (improving; Rev +36%)
     "ABOVE AL1 — NEAR ACCUMULATE. Macro Fib (2022 low $3.48 → ATH $99.58): "
     "AL1=$62.87 | AL2=$51.53 | AL3=$40.18 | 78.6%=$23.99. "
     "Current $65.94 just ABOVE AL1 ($62.87) — AL1 entry zone very close. "
     "DoD/NSSL contracts. D200=$57.23; current above D200. W20=$66.86 — price at W20 EMA.",
     STK, None),

    ("Intuitive Machines", "LUNR", "USD", 24.41,
     6.135, 23.315,
     None, None, 0.64, -0.38,  # D/E=0.64 (space debt); net margin approx -38% (pre-profit, NASA-dependent)
     "ABOVE AL1 — WATCH. Macro Fib (inception low $6.135 → ATH $23.315): "
     "AL1=$16.75 | AL2=$14.72 | AL3=$12.70. "
     "Current $17.52 above AL1. Set alert at AL1 ($16.75). NASA CLPS contractor.",
     STK, None),

    ("LONG-TERM FUND HOLDINGS", "", "", None, None, None, None, None, None, None, "", S, None),

    ("Vanguard FTSE All-World Acc", "VWRP.L", "GBP", 130.00,
     75.24, 140.92,
     None, None, None, None,
     "AL1 HIT — ACCUMULATE. Macro Fib (2022 low £75.24 → ATH £140.92): "
     "AL1=£115.30 | AL2=£108.08 | AL3=£100.87 | 78.6%=£89.78. "
     "Current £123.58 BELOW macro AL1? No — current £123.58 above AL1 (£115.30). "
     "3 tranches bought avg ~£126-128. 4th tranche planning. "
     "D200=£122.60; current £123.58 just above D200 — very close to D200 support. "
     "W20=£126.06 / W50=£121.11 — price between W20 and W50.",
     SIPP_T, None),

    ("VanEck Semiconductor", "SMGB.L", "GBP", 58.35,
     13.79, 55.76,
     None, None, None, None,
     "ABOVE AL1. Macro Fib (2022 low £13.79 → ATH £55.76): "
     "AL1=£39.69 | AL2=£34.78 | AL3=£29.86. "
     "Current £50.65 well above AL1 — near ATH. 10% SIPP allocation. "
     "D200=£43.66; current above D200. W20=£49.15 / W50=£43.07.",
     SIPP_T, None),

    ("Vanguard S&P 500 UCITS ETF", "VUSA.L", "GBP", 96.51,
     53.80, 107.01,
     None, None, None, None,
     "ABOVE AL1 — WATCH / HOLD. Macro Fib (Jun 2022 low £53.80 → ATH £107.01): "
     "AL1=£86.68 | AL2=£80.41 | AL3=£74.13 | 78.6%=£65.19. "
     "Current £96.51 well above AL1 (£86.68) — not at entry. SIPP core holding. "
     "VUSA tracks S&P 500. Vanguard, 0.07% OCF. Unhedged — GBP/USD affects returns in GBP. "
     "Accumulate on pullbacks: AL1 (£86.68) = strong add; AL2 (£80.41) = back up truck; AL3 (£74.13) = conviction buy. "
     "Apr 2025 tariff crash low: £69.89 intraweek — 78.6% Fib £65.19 was not hit. "
     "D200=£94.06; current above D200 — bullish. "
     "W20=£94.95 / W50=£92.41 — price above both weekly EMAs, bullish cross. "
     "M20=£89.19 / M50=£77.96 — strongly bullish long-term structure.",
     SIPP_T, None),

    # ── Growth watchlist ──────────────────────────────────────────────────────
    ("GROWTH WATCHLIST — thesis positions", "", "", None, None, None, None, None, None, None, "", S, None),

    ("NVIDIA Corp", "NVDA", "USD", 189.31,
     10.80, 212.17,
     36.0, 16.0, 0.40, 0.56,
     "ABOVE AL1 — ABOVE MACRO ENTRY. "
     "Macro Fib (2022 low $10.80 → ATH $212.17): AL1=$135.24 | AL2=$111.49 | AL3=$87.68. "
     "Current $166.60 above macro AL1 ($135.24) — not at macro entry yet. "
     "Fwd P/E 16x. Rev +73%. "
     "D200=$91.67; price well above D200 — note D200 is still rising, not a breakdown. "
     "W20=$179.82 / W50=$168.75 — current between W20 and W50, bearish crossover risk.",
     ISA_T, None),

    ("Amazon.com", "AMZN", "USD", 239.89,
     81.69, 258.60,
     29.0, 23.0, 0.43, 0.11,
     "ABOVE AL1 — NEAR ENTRY. "
     "Macro Fib (2022 low $81.69 → ATH $258.60): AL1=$191.07 | AL2=$170.15 | AL3=$149.22. "
     "Current $198.42 above AL1 ($191.07) — holding above macro accumulate zone. "
     "AL1 ($191.07) = next key level to watch. AWS + ads + Prime. "
     "D200=$167.90; price well above D200. W20=$217.70 / W50=$217.23 — price below both.",
     ISA_T, None),

    ("Microsoft Corp", "MSFT", "USD", 384.37,
     207.39, 552.24,
     23.0, 20.0, 0.43, 0.39,
     "AL2 HIT — STRONG BUY. Macro Fib (2022 low $207.39 → ATH $552.24): "
     "AL1=$420.37 | AL2=$379.82 | AL3=$339.24 | 78.6%=$281.27. "
     "Current $365.97 between AL2 ($379.82) and AL3 ($339.24) — STRONG BUY ZONE. "
     "Azure AI + Office 365 + Copilot. Rev +16.7%, earnings +59.8%. Fwd P/E 20x. Net margin 39%. "
     "D200=$373.19; current near but below D200 — bearish signal. "
     "W20=$429.31 / W50=$447.79 — price well below weekly EMAs.",
     ISA_T, None),

    ("Oracle Corporation", "ORCL", "USD", 155.62,
     58.11, 344.21,
     25.0, 17.0, None, 0.253,   # D/E N/A — negative equity from buybacks; $162B gross debt noted below
     "AL3 HIT — BACK UP TRUCK. Macro Fib (2022 low $58.11 → ATH $344.21 Sep 2025): "
     "AL1=$234.92 | AL2=$201.16 | AL3=$167.40 | 78.6%=$119.34. "
     "Current $155.62 — BELOW AL3 ($167.40), above 78.6% ($119.34). "
     "4th Hyperscaler alongside AWS/Azure/GCP. OCI (Oracle Cloud Infrastructure): Rev +52% YoY. "
     "Stargate AI: $500B US AI infrastructure JV (OpenAI/SoftBank/ORCL) — OCI designated primary cloud. "
     "Net margin 25%. Fwd P/E 17x. Rev +22% YoY (FY2026). Free cash flow positive. "
     "DEBT NOTE: $162B gross debt / $39B cash = $123B net debt. Shareholders equity NEGATIVE due to buybacks — "
     "D/E ratio not meaningful; leverage is structural, not distress. "
     "D200=$216.27; price well below D200 — bear regime. W20=$169.63 / W50=$184.14 — bearish. "
     "M20=$173.18; M50=$142.38 — approaching M50 support. Set accumulate alert at AL3 ($167.40).",
     ISA_T, None),

    ("Meta Platforms", "META", "USD", 634.53,
     87.40, 794.38,
     25.0, 17.0, 0.33, 0.30,
     "NEAR AL1 — APPROACHING ACCUMULATE. Macro Fib (2022 low $87.40 → ATH $794.38): "
     "AL1=$524.31 | AL2=$440.89 | AL3=$357.47 | 78.6%=$238.69. "
     "Current $547.54 just ABOVE macro AL1 ($524.31) — aka 61.8% retracement on TradingView. "
     "Monthly chart: price sitting precisely on 61.8% retrace level — institutional support zone. "
     "M50 EMA=$485.12 = next stop below AL1. D200=$436.05 = below M50. "
     "Set alert at AL1 ($524.31). Strong buy on any close below $524. "
     "Facebook/IG/WhatsApp + AI. Rev +23.8%. Fwd P/E 17x. Net margin 30%.",
     ISA_T, None),

    ("Alphabet (Google)", "GOOGL", "USD", 321.31,
     82.66, 348.75,
     27.0, 22.0, 0.22, 0.33,
     "ABOVE AL1. Macro Fib (2022 low $82.66 → ATH $348.75): "
     "AL1=$246.04 | AL2=$215.71 | AL3=$185.35. "
     "Current $280.92 above AL1 ($246.04) — above accumulate zone. "
     "Set alert at AL1 ($246.04). GSearch + YouTube + GCP + Waymo. "
     "D200=$163.87; price well above D200. W20=$297.37 / W50=$261.30.",
     ISA_T, None),

    ("Netflix", "NFLX", "USD", 103.16,
     16.27, 134.12,
     36.0, 24.0, 0.59, 0.24,
     "ABOVE AL1. Macro Fib (2022 low $16.27 → ATH $134.12): "
     "AL1=$89.04 | AL2=$75.20 | AL3=$61.35. "
     "Current $93.32 just above AL1 ($89.04) — very close to accumulate zone. "
     "Set alert at AL1 ($89.04). Streaming monopoly + ads tier + games. "
     "D200=$107.65; price below D200 — bearish. W20=$94.70 / W50=$100.58.",
     ISA_T, None),

    ("Apple Inc.", "AAPL", "USD", 259.20,
     123.88, 288.35,
     33.0, 28.0, None, 0.270,   # D/E N/A — negative equity from buybacks; net debt $23.6B trivial vs $100B+ FCF
     "ABOVE AL1 — WAIT FOR PULLBACK. Macro Fib (2022 low $123.88 → ATH $288.35 Dec 2025): "
     "AL1=$225.52 | AL2=$206.12 | AL3=$186.71 | 78.6%=$159.08. "
     "Current $260.52 well above AL1 ($225.52) — only 10.7% upside to ATH. Not a buy here. "
     "THESIS: iPhone installed base 2.2B devices. Services ($100B+ ARR, 30%+ margin) now >25% revenue. "
     "Apple Intelligence (AI features) = upgrade supercycle catalyst. "
     "TARIFF RISK: ~85% iPhone assembly in China — US-China tariffs are structural headwind; "
     "India/Vietnam capacity buildout partially mitigates but multi-year transition. "
     "Net margin 27%. Fwd P/E 28x — premium vs sector, justified by ecosystem lock-in + buybacks. "
     "$110B+ buybacks/year = structural price floor. "
     "D200=$250.31; price above D200 — bullish. W20=$257.26 / W50=$244.91 / M20=$239.38. "
     "Set accumulate alert at AL1 ($225.52); strong buy at AL2 ($206.12).",
     ISA_T, None),


    ("Tesla, Inc.", "TSLA", "USD", 352.42,
     101.81, 498.83,
     326.0, 128.84, 0.18, 0.040,
     "AL1 ZONE — CAUTIOUS ACCUMULATE. Macro Fib (2022 low $101.81 → ATH $498.83): "
     "AL1=$347.17 | AL2=$300.32 | AL3=$253.47 | 78.6%=$186.77. "
     "Current $352.08 just ABOVE AL1 ($347.17) — in the accumulate zone. "
     "THESIS: FSD/Full-Self-Driving (Robotaxi launch 2026) + Optimus humanoid robot + Energy storage (Megapack) growth. "
     "HEADWINDS: Revenue -3.1% YoY; earnings -60.6% YoY; Elon Musk brand risk (DOGE) impacting sales Europe/China. "
     "VALUATION: TTM P/E 326x / Fwd P/E 129x — extreme premium; priced for autonomy success. Net cash +$29B ($44B cash vs $15B debt). "
     "D200=$397.35; price well below D200 — bearish regime. W20=$393.10 / W50=$380.53 — price below both weekly EMAs. "
     "M20=$351.53 — price just at M20 (key support). M50=$336.30. "
     "Set strong buy at AL2 ($300.32); back-up-truck at AL3 ($253.47).",
     ISA_T, None),

    ("Intel Corporation", "INTC", "USD", 65.18,
     23.40, 68.49,
     None, 63.95, 0.49, -0.00505,
     "ABOVE AL1 — APPROACHING CYCLE ATH. NOT AN ENTRY. Macro Fib (Oct 2022 bear low $23.40 → Feb 2021 ATH $68.49): "
     "AL1=$51.27 | AL2=$45.94 | AL3=$40.62 | 78.6%=$33.05. "
     "Current $65.18 well above AL1 ($51.27) — do NOT chase here. "
     "THESIS: Speculative turnaround on Intel Foundry Services (IFS) + CHIPS Act. "
     "New CEO Lip-Bu Tan (Mar 2025) restructuring aggressively — cost cuts, factory refocus. "
     "US CHIPS Act: $8.5B grants + $11B loans approved; critical to domestic semiconductor sovereignty. "
     "Apr 2025 tariff crash low: $18.13 (intraweek). Stock has since tripled on deal/catalyst flow. "
     "RISKS: Currently LOSS-MAKING (net margin -0.5%); foundry ramp takes 2–3 years; AMD/TSMC competition structural. "
     "VALUATION: Fwd P/E 64x — expensive for a turnaround; D/E 0.49 — manageable debt. "
     "D200=$36.41; price massively above D200 — extended, mean-reversion risk is HIGH. "
     "W20=$46.06 / W50=$36.82 — price far above both weekly EMAs. "
     "M20=$36.03 / M50=$31.91 — long-term structure recovering. "
     "WAIT for deep pullback: AL1 ($51.27) = initial watch; AL2 ($45.94) = add; AL3 ($40.62) = conviction buy. "
     "Set GTC alert at $51.27 (AL1). Only enter on confirmed pullback — avoid FOMO at ATH.",
     ISA_T, None),

    ("XRP", "XRP-USD", "USD", 1.3766,
     0.29, 3.84,
     None, None, None, None,
     "BELOW AL3 — DEEP VALUE. Macro Fib (2022 low $0.29 → all-time ATH $3.84 Jan 2018): "
     "AL1=$2.48 | AL2=$2.07 | AL3=$1.64 | 78.6%=$1.05. "
     "Current $1.36 between AL3 ($1.64) and 78.6% ($1.05) — extreme value zone. "
     "GENIUS Act: RLUSD stablecoin legitimised. Ripple regulatory clarity = bullish long-term. "
     "D200=$2.06; price well below D200. "
     "W20=$1.70 / W50=$2.03 — bearish. Add below $1.64.",
     STK, None),

    ("Bitmine Immersion Tech", "BMNR", "USD", 21.51,
     3.92, 160.95,
     None, 21.0, 0.00, -3.0,  # Deeply loss-making ETH treasury; staking revenue << operating costs
     "#1 Ethereum treasury (4.6M ETH staked). Launched Jun 2025 — 9mo history, no 2022 macro anchor. "
     "Down 87% from $161 high. ETH at $2,059 (below macro AL3 $2,446). "
     "GENIUS Act: staking yield NOT a stablecoin yield — thesis intact. "
     "D200=$36.63 / W20=$26.31 — price below all MAs. Do not buy until crypto floor confirmed. "
     "Manual levels: $30/$20/$10.",
     MAN, (30.00, 20.00, 10.00)),

    ("Ethereum (Reference)", "ETH-USD", "USD", 2369.90,
     896.11, 4953.73,
     None, None, None, None,
     "BELOW AL3 — DEEP VALUE (REFERENCE ONLY). Macro Fib (2022 low $896 → ATH $4,954): "
     "AL1=$3,404 | AL2=$2,925 | AL3=$2,446 | 78.6%=$1,764. "
     "Current $2,370 between AL3 ($2,446) and 78.6% ($1,764) — approaching AL3. "
     "GENIUS Act tailwind: ETH is the primary stablecoin settlement layer ($1T+ daily volume). "
     "D200=$3,093; price well below D200. W20=$2,511 / M50=$2,503 — bearish structure. "
     "Not held. Monitor for entry at AL3 ($2,446) reclaim.",
     STK, None),

    # ── BTC REFERENCE ─────────────────────────────────────────────────────────
    ("BTC REFERENCE LEVELS", "", "", None, None, None, None, None, None, None, "", S, None),

    ("Bitcoin", "BTC", "USD", 74442.23,
     15599.05, 126198.07,
     None, None, None, None,
     "AL2 HIT — STRONG BUY (MACRO). Macro Fib (2022 low $15,599 → ATH $126,198): "
     "AL1=$83,949 | AL2=$70,899 | AL3=$57,848 | 78.6%=$39,267. "
     "Current $74,442 — ABOVE AL2 ($70,899), reclaimed. AL2 reclaim = bullish signal. "
     "D200=$91,315; price well below D200 — BEAR REGIME CONFIRMED. "
     "W20=$79,354 / W50=$88,689 / M20=$83,486 / M50=$65,485 — all above current price. "
     "ISA rotation thesis RE-ENGAGED — AL2 ($70,899) reclaimed. Next resistance: AL1 ($83,949). "
     "Next support: AL3=$57,848. GENIUS Act OCC rules Mar 2026 — market digesting.",
     STK, None),

    # ── PRE-IPO ───────────────────────────────────────────────────────────────
    ("PRE-IPO ACCESS VEHICLES — LSE investment trusts: SpaceX, Anthropic, Bytedance. ISA-eligible.", "", "", None, None, None, None, None, None, None, "", S, None),

    ("Scottish Mortgage Inv Trust", "SMT.L", "GBp", 1373.16,
     670.47, 1568.18,
     None, None, None, None,
     "AL1 HIT — ACCUMULATE. Macro Fib (2022 low 670p → ATH 1568p Nov 2021): "
     "AL1=1225p | AL2=1119p | AL3=1013p | 78.6%=862p. "
     "Current 1185.8p BELOW AL1 (1225p) — accumulate zone. SpaceX top holding (~15% NAV). "
     "D200=1130p; current above D200. W20=1175p / W50=1116p — price just above both. "
     "SpaceX/Anthropic IPOs = major NAV catalysts. Trades at discount to NAV.",
     ISA_T, None),

    ("Schiehallion Fund", "MNTN.L", "USD", 1.92,
     0.83, 3.08,
     None, None, None, None,
     "BELOW AL3 — DEEP VALUE. Macro Fib (2022 low $0.83 → ATH $3.08 Nov 2021): "
     "AL1=$2.22 | AL2=$1.96 | AL3=$1.69 | 78.6%=$1.31. "
     "Current $1.72 between AL3 ($1.69) and AL2 ($1.96) — deep value zone. "
     "D200=$1.35; current above D200. W20=$1.58 / M50=$1.25 — mixed MA picture. "
     "Pure-play SpaceX+Anthropic. Higher upside but less liquid than SMT.L.",
     ISA_T, None),

    # ── ASYMMETRIC GROWTH BASKET ───────────────────────────────────────────────
    ("ASYMMETRIC GROWTH BASKET — High-conviction speculative positions: AI software, biotech/genomics, quantum computing, defence tech, space. High risk / high reward. Set Fib anchors manually once conviction targets confirmed.", "", "", None, None, None, None, None, None, None, "", S, None),

    ("Schrödinger", "SDGR", "USD", None, 14.03, 109.64,
     None, None, 0.00, -0.50,
     "SPECULATIVE — AI-DRIVEN DRUG DISCOVERY. Physics-based computational platform for drug discovery. "
     "Revenue from software + drug pipeline milestones. Pre-profit; net margin ~-50%. "
     "Partnerships: BMS, GSK, Eli Lilly. Revenue growing ~20% YoY. "
     "Macro Fib anchors: 2022 bear low $14.03 → cycle ATH $109.64. Set alert at AL1 ($75.93) and AL2 ($61.83). "
     "Set Fib alerts manually. Cash-rich balance sheet (~$500M). No debt.",
     STK, None),

    ("Penguin Solutions", "PENG", "USD", None, None, None,
     None, None, None, None,
     "SPECULATIVE — HPC/AI MEMORY & STORAGE. Formerly SMART Global Holdings (SGH). "
     "Designs memory modules and HPC solutions for AI/data centre workloads. "
     "Revenue ~$400M. Transitioning to higher-margin AI infrastructure solutions. "
     "Fib anchors: set manually once price history establishes under PENG ticker. Set alert levels manually.",
     STK, None),

    ("UiPath", "PATH", "USD", None, 10.55, 90.00,
     None, None, 0.00, -0.05,
     "SPECULATIVE — AGENTIC AI / AUTOMATION. Enterprise automation platform: RPA + AI agents. "
     "~$1.4B revenue run-rate, growing ~10% YoY. Microsoft partnership. SAP integration. "
     "Path to profitability: FCF positive. No debt. ~$1.8B cash. "
     "Macro Fib anchors: 2022 bear low $10.55 → IPO ATH $90.00. "
     "AL1=$56.73 | AL2=$50.27 | AL3=$43.81 | 78.6%=$33.26. "
     "Activist pressure → cost discipline. Watch for AI agent monetisation.",
     STK, None),

    ("SentinelOne", "S", "USD", None, 12.43, 77.86,
     None, None, 0.00, -0.15,
     "SPECULATIVE — AI-NATIVE CYBERSECURITY. Singularity platform: XDR + CNAPP + AI-SOC. "
     "Revenue ~$800M growing ~30% YoY. Competing vs CrowdStrike post-CSST outage (Jul 2024 gained share). "
     "Pre-profit; improving margins. No debt. ~$1.1B cash. "
     "Macro Fib anchors: 2022 bear low $12.43 → ATH $77.86. "
     "AL1=$50.93 | AL2=$45.14 | AL3=$39.36 | 78.6%=$30.04. "
     "Thesis: AI SOC automation + Purple AI analyst assistant driving upsell.",
     STK, None),

    ("Flywire", "FLYW", "USD", None, 12.63, 48.66,
     None, None, 0.00, 0.02,
     "SPECULATIVE — GLOBAL PAYMENTS PLATFORM. Vertical-specific payments for healthcare, education, travel. "
     "Revenue ~$440M growing ~25% YoY. Near-breakeven FCF. No debt. "
     "Macro Fib anchors: 2022 bear low $12.63 → ATH $48.66. "
     "AL1=$34.94 | AL2=$30.64 | AL3=$26.35 | 78.6%=$20.39. "
     "Thesis: global tuition/healthcare payment flows with high switching costs.",
     STK, None),

    ("American Superconductor", "AMSC", "USD", None, 6.14, None,
     None, None, 0.00, -0.10,
     "SPECULATIVE — POWER ELECTRONICS / GRID + DEFENCE. Wind turbine electrical systems + grid solutions. "
     "Also naval defence: degaussing systems for US Navy. Revenue ~$150M. Pre-profit. "
     "NOTE: stock surged past prior ATH (~$40) on AI grid demand — set new cycle ATH manually once confirmed. "
     "2022 bear low $6.14 still valid as lower anchor. "
     "Thesis: power grid modernisation + defence superconductor tech.",
     STK, None),

    ("PROCEPT BioRobotics", "PRCT", "USD", None, 19.85, 132.98,
     None, None, 0.00, -0.40,
     "SPECULATIVE — SURGICAL ROBOTICS (UROLOGY). AquaBeam robotic system for BPH treatment. "
     "Revenue ~$300M growing ~60% YoY. Pre-profit but rapid commercialisation. "
     "Macro Fib anchors: 2022 IPO low $19.85 → ATH $132.98. "
     "AL1=$89.75 | AL2=$76.42 | AL3=$63.08 | 78.6%=$43.81. "
     "Thesis: only FDA-cleared robotic waterjet BPH system. TAM $3B+. Urology robotics disruptor.",
     STK, None),

    ("Pagaya Technologies", "PGY", "USD", None, None, None,
     None, None, None, None,
     "SPECULATIVE — AI CREDIT MARKETPLACE. AI-driven credit decisioning embedded in lender workflows. "
     "Partners: Ally, SoFi, US Bank. Revenue ~$900M but high volatility. Pre-profit. "
     "SPAC history → complex capital structure. Watch dilution risk. "
     "Set Fib anchors manually once stable price regime established. Monitor $2-5 range for accumulation.",
     STK, None),

    ("Intellia Therapeutics", "NTLA", "USD", None, 20.29, 222.12,
     None, None, 0.00, -1.50,
     "SPECULATIVE — IN VIVO CRISPR GENE EDITING. Systemic CRISPR/Cas9 delivery platform. "
     "Lead: NTLA-2001 (transthyretin amyloidosis — Phase 3). NTLA-2002 (HAE — Phase 2). "
     "Cash runway ~2 years. No revenue (clinical stage). Partnered with Regeneron. "
     "Macro Fib anchors: 2022 bear low $20.29 → ATH $222.12. "
     "AL1=$137.56 | AL2=$121.21 | AL3=$104.84 | 78.6%=$76.10. "
     "Thesis: first systemic in vivo CRISPR therapy → platform value. Binary binary on Phase 3 data.",
     STK, None),

    ("Beam Therapeutics", "BEAM", "USD", None, 22.25, 145.80,
     None, None, 0.00, -1.80,
     "SPECULATIVE — BASE EDITING (NEXT-GEN CRISPR). Precision base editing without double-strand DNA breaks. "
     "Pipeline: BEAM-101 (sickle cell/beta-thal), BEAM-302 (AAT deficiency). Clinical stage. "
     "Cash ~$800M (extended runway). Partnered with Pfizer ($300M deal). "
     "Macro Fib anchors: 2022 bear low $22.25 → ATH $145.80. "
     "AL1=$98.59 | AL2=$84.02 | AL3=$69.47 | 78.6%=$48.89. "
     "Thesis: base editing is cleaner than standard CRISPR — potential platform licensing value.",
     STK, None),

    ("Absci", "ABSI", "USD", None, 3.15, 29.80,
     None, None, 0.00, -1.20,
     "SPECULATIVE — GENERATIVE AI FOR DRUG DISCOVERY. AI protein design platform (generative biology). "
     "Drug creation using AI + synthetic biology. Pre-revenue (platform licensing stage). "
     "Partnerships: AstraZeneca, Merck. Cash ~$250M. "
     "Macro Fib anchors: 2022 bear low $3.15 → ATH $29.80. "
     "AL1=$19.63 | AL2=$16.47 | AL3=$13.32 | 78.6%=$8.85. "
     "Thesis: AI-designed antibodies at scale — zero-shot drug design platform. Long-duration bet.",
     STK, None),

    ("BlackSky Technology", "BKSY", "USD", None, None, None,
     None, None, None, -0.60,
     "SPECULATIVE — AI EARTH OBSERVATION INTELLIGENCE. Sub-daily satellite imagery + AI analytics. "
     "Government/defence customers (NRO, NGA). Revenue ~$100M growing ~20% YoY. "
     "NOTE: BKSY underwent 1:10 reverse split (May 2023) — set Fib anchors manually in post-split terms. "
     "Pre-split low $1.20 → post-split ~$12. Set new ATH manually from post-split price history. "
     "Thesis: real-time intelligence from space — AI-fusion of imagery + signals data.",
     STK, None),

    ("Red Cat Holdings", "RCAT", "USD", None, 1.52, 14.85,
     None, None, 0.00, -0.80,
     "SPECULATIVE — MILITARY DRONE SYSTEMS. Small UAS (FPV drones) for US DoD. "
     "Black Widow and Fleep drones qualified for US Army SRR programme. "
     "Revenue ~$35M growing rapidly. NDAA compliant (no Chinese components). "
     "Macro Fib anchors: 2022 bear low $1.52 → ATH $14.85. "
     "AL1=$9.63 | AL2=$8.18 | AL3=$6.72 | 78.6%=$4.70. "
     "Thesis: US-made defence drone supplier. Drone warfare structural demand. Small float.",
     STK, None),

    ("Planet Labs", "PL", "USD", None, None, None,
     None, None, None, -0.70,
     "SPECULATIVE — DAILY EARTH IMAGING + GEOSPATIAL ANALYTICS. Dove/SkySat constellation. "
     "Daily global imaging coverage. Revenue ~$240M growing ~10% YoY. Pre-profit. "
     "NOTE: verify price history for reverse split adjustments before setting Fib anchors. "
     "Set bear low and ATH manually from post-split adjusted history. "
     "Thesis: only company offering daily global imagery at scale. Data-as-a-service moat.",
     STK, None),

    ("Denison Mines", "DNN", "USD", None, 0.96, 3.76,
     None, None, 0.20, None,
     "SPECULATIVE — URANIUM DEVELOPER (ATHABASCA). Wheeler River project (Phoenix ISR + Gryphon). "
     "Pre-production; Phoenix ISR could be lowest-cost uranium mine globally. EA approved. "
     "23.9% stake in Orano Canada joint venture. No revenue (developer). "
     "Macro Fib anchors: 2022 bear low $0.96 → ATH $3.76. "
     "AL1=$2.69 | AL2=$2.36 | AL3=$2.04 | 78.6%=$1.60. "
     "Thesis: ISR uranium mining at Wheeler River = structural supply to nuclear renaissance.",
     STK, None),

    ("NanoXplore", "GRA", "USD", None, None, None,
     None, None, None, None,
     "SPECULATIVE — GRAPHENE MATERIALS. Mass-production graphene powder for EV batteries, composites. "
     "Trades on TSX Venture Exchange (symbol: GRA). Canadian company (revenues in CAD). "
     "Revenue ~CAD $50M growing. Pre-profit. Customers: GM, General Dynamics. "
     "Thesis: graphene battery enhancement + lightweighting composites for EV/aerospace. "
     "Set Fib anchors manually in CAD. Note: live price may require GRA.TO in yfinance.",
     STK, None),

    ("QuantumScape", "QS", "USD", None, 4.78, 131.67,
     None, None, None, None,
     "SPECULATIVE — SOLID-STATE BATTERY (EV). Lithium-metal solid-state separator technology. "
     "Pre-revenue; pilot production. VW backed. CEO Jagdeep Singh. "
     "Macro Fib anchors: SPAC low $4.78 → ATH $131.67. "
     "AL1=$86.30 | AL2=$68.23 | AL3=$50.17 | 78.6%=$25.73. "
     "Thesis: solid-state battery is holy grail for EVs — if commercialised, transformational. "
     "Very long duration / binary. Watch 2026 production milestones with VW.",
     STK, None),

    ("IonQ", "IONQ", "USD", None, 3.25, 54.89,
     None, None, 0.00, None,
     "SPECULATIVE — TRAPPED-ION QUANTUM COMPUTING. Hardware + cloud access (AWS/Azure/Google Cloud). "
     "Revenue ~$40M growing ~100% YoY. Pre-profit. #AQ metric (algorithmic qubits). "
     "Macro Fib anchors: 2022 bear low $3.25 → ATH $54.89. "
     "AL1=$35.75 | AL2=$29.07 | AL3=$22.38 | 78.6%=$12.87. "
     "Thesis: trapped-ion has lower error rates than superconducting — potential enterprise advantage. "
     "US DoD and AFRL contracts validate defence quantum use case.",
     STK, None),

    ("D-Wave Quantum", "QBTS", "USD", None, 0.88, 21.42,
     None, None, 0.00, None,
     "SPECULATIVE — QUANTUM ANNEALING COMPUTING. Advantage system for optimisation problems. "
     "Revenue ~$10M. Only commercially deployed quantum system at scale. "
     "Macro Fib anchors: 2022/2023 low $0.88 → ATH $21.42. "
     "AL1=$13.71 | AL2=$11.15 | AL3=$8.59 | 78.6%=$4.60. "
     "Thesis: optimisation use cases (logistics, pharma, finance) deployable today. "
     "First-mover in commercial quantum — but annealing vs gate-based model debate ongoing.",
     STK, None),

    ("Rigetti Computing", "RGTI", "USD", None, 0.77, 29.65,
     None, None, 0.00, None,
     "SPECULATIVE — SUPERCONDUCTING QUANTUM COMPUTING. Novera QPU + cloud Quantum Computing Service. "
     "Revenue ~$10M. SPAC merged 2022. US-made superconducting chips. "
     "Macro Fib anchors: SPAC low $0.77 → ATH $29.65. "
     "AL1=$19.38 | AL2=$15.21 | AL3=$11.05 | 78.6%=$4.33. "
     "Thesis: US domestic quantum hardware supplier — NDAA / national security angle. "
     "84-qubit Ankaa-3 system. Watch for 100+ qubit milestone.",
     STK, None),

    # ── IPO WATCHLIST ──────────────────────────────────────────────────────────
    ("IPO WATCHLIST — PRE-IPO MONITORING  |  Accessible now via SMT.L / MNTN.L.", "", "", None, None, None, None, None, None, None, "", S, None),

    ("SpaceX", "TBC", "USD", None, None, None,
     None, None, None, None,
     "PRE-IPO. Est. valuation ~$1.5T. Expected IPO mid-to-late 2026. "
     "Starship, Starlink, Falcon 9. ACCESSIBLE NOW via SMT.L (15% NAV) and MNTN.L (top holding).",
     STK, None),

    ("OpenAI", "TBC", "USD", None, None, None,
     None, None, None, None,
     "PRE-IPO. Est. valuation ~$1T (Jan 2026 funding). Expected IPO 2027. "
     "ChatGPT >300M weekly users. INDIRECT EXPOSURE via MSFT (ISA).",
     STK, None),

    ("Anthropic", "TBC", "USD", None, None, None,
     None, None, None, None,
     "PRE-IPO. Est. valuation ~$350B (Nov 2025). Expected IPO 2026 possible. "
     "Google + Amazon backers. Claude family. INDIRECT EXPOSURE via AMZN (ISA) + MNTN.L.",
     STK, None),
]

# Apply live prices and fundamentals over hardcoded fallbacks
if live_prices or live_fundamentals:
    _rd = []
    for _rec in rows_data:
        _a, _t, _c, _cur, _lo, _hi, _ttm, _fwd, _de, _mg, _nt, _rt, _man = _rec
        # Price — override when live data exists; None-price entries not in _FETCH
        # (e.g. CRCL, TBC) won't appear in live_prices so are safely excluded
        if _t in live_prices:
            _cur = live_prices[_t]
        # Fundamentals — only override when live value is not None (preserves ETF/N/A intent)
        if _t in live_fundamentals:
            _lf = live_fundamentals[_t]
            if _lf[0] is not None: _ttm = _lf[0]
            if _lf[1] is not None: _fwd = _lf[1]
            if _lf[2] is not None: _de  = _lf[2]
            if _lf[3] is not None: _mg  = _lf[3]
        _rd.append((_a, _t, _c, _cur, _lo, _hi, _ttm, _fwd, _de, _mg, _nt, _rt, _man))
    rows_data = _rd

# ── Fetch and inject --add tickers ────────────────────────────────────────────
def _fetch_new_ticker_row(tkr):
    """Fetch live yfinance data for a dynamically added ticker.

    Returns a rows_data-compatible 13-tuple, or None on failure.
    Also updates live_prices / live_ma / live_fundamentals / live_analyst dicts
    so the row is treated identically to a hardcoded entry.
    """
    try:
        import yfinance as _yf

        _yt   = _yf.Ticker(tkr)
        _info = _yt.info
        _name = _info.get("longName") or _info.get("shortName") or tkr
        _ccy  = _info.get("currency", "USD")

        # 5-year daily history for MAs, ATH, and 2022 bear low
        _hist = _yt.history(period="5y", interval="1d")
        if _hist.empty:
            print(f"[--add {tkr}] No price history — skipping.")
            return None

        _close = _hist["Close"].dropna()
        _cur   = round(float(_close.iloc[-1]), 2)

        # ATH: highest close in available 5Y history
        _ath   = round(float(_close.max()), 2)

        # 2022 bear market low: lowest intraday low across Oct–Dec 2022
        _bear_window = _hist.loc["2022-09-01":"2022-12-31"]
        _lo = (round(float(_bear_window["Low"].min()), 2)
               if not _bear_window.empty
               else round(float(_hist["Low"].min()), 2))

        # D200 SMA
        _d200 = (round(float(_close.rolling(200).mean().iloc[-1]), 2)
                 if len(_close) >= 200 else None)

        # Weekly EMA — resample daily close to Friday close
        _wkly = _close.resample("W-FRI").last().dropna()
        _w20  = (round(float(_wkly.ewm(span=20, adjust=False).mean().iloc[-1]), 2)
                 if len(_wkly) >= 20 else None)
        _w50  = (round(float(_wkly.ewm(span=50, adjust=False).mean().iloc[-1]), 2)
                 if len(_wkly) >= 50 else None)

        # Monthly EMA — resample to month-end
        _mthly = _close.resample("ME").last().dropna()
        _m20   = (round(float(_mthly.ewm(span=20, adjust=False).mean().iloc[-1]), 2)
                  if len(_mthly) >= 20 else None)
        _m50   = (round(float(_mthly.ewm(span=50, adjust=False).mean().iloc[-1]), 2)
                  if len(_mthly) >= 50 else None)

        # Fundamentals
        _ttm = _info.get("trailingPE")
        if _ttm and _ttm <= 0:
            _ttm = None
        if _ttm:
            _ttm = round(_ttm, 1)
        _fpe = _info.get("forwardPE")
        if _fpe:
            _fpe = round(_fpe, 1)
        _de_raw = _info.get("debtToEquity")
        _de = round(_de_raw / 100, 2) if _de_raw is not None else None
        _mg = _info.get("profitMargins")
        if _mg is not None:
            _mg = round(_mg, 4)

        # Analyst consensus
        _rec_map = {
            "strong_buy": "Strong Buy", "strongbuy": "Strong Buy",
            "buy": "Buy", "hold": "Hold",
            "underperform": "Sell", "sell": "Sell", "strong_sell": "Sell",
        }
        _cons  = _rec_map.get(_info.get("recommendationKey", "").lower(), "Hold")
        _n_ana = int(_info.get("numberOfAnalystOpinions") or 0)
        _tgt   = _info.get("targetMeanPrice")
        _tgt   = round(float(_tgt), 2) if _tgt else None

        # Register in live dicts so the run header reflects the fetch
        live_prices[tkr]       = _cur
        live_ma[tkr]           = (_w20, _w50, _m20, _m50, _d200)
        live_fundamentals[tkr] = (_ttm, _fpe, _de, _mg)
        live_analyst[tkr]      = (_cons, _n_ana, _tgt)

        # Compute Fib levels for the auto-generated note
        def _fl(hi, lo, r):
            return round(hi - (hi - lo) * r, 2)

        _al1 = _fl(_ath, _lo, 0.382)
        _al2 = _fl(_ath, _lo, 0.500)
        _al3 = _fl(_ath, _lo, 0.618)
        _d4  = _fl(_ath, _lo, 0.786)

        # Zone descriptor
        if _cur < _d4:
            _zone = f"BELOW 78.6% ({_d4}) — DEEPEST VALUE."
        elif _cur < _al3:
            _zone = f"AL3 HIT ({_al3}) — BACK UP TRUCK."
        elif _cur < _al2:
            _zone = f"AL2 HIT ({_al2}) — STRONG BUY."
        elif _cur < _al1:
            _zone = f"AL1 HIT ({_al1}) — ACCUMULATE."
        else:
            _zone = f"ABOVE AL1 ({_al1}) — WATCHING."

        _note = (
            f"{_zone} "
            f"Macro Fib (2022 low {_lo} -> ATH {_ath}): "
            f"AL1={_al1} | AL2={_al2} | AL3={_al3} | 78.6%={_d4}. "
            f"D200={_d200} | W20={_w20} / W50={_w50} / M20={_m20} / M50={_m50}. "
            f"TTM P/E={_ttm or 'N/A'} | Fwd P/E={_fpe or 'N/A'} | "
            f"D/E={_de or 'N/A'} | Net margin={round(_mg * 100, 1) if _mg else 'N/A'}%. "
            f"[Auto-added via --add {tkr}]"
        )
        print(f"[--add {tkr}] Fetched OK — cur={_cur}, ATH={_ath}, 2022 low={_lo}, "
              f"AL1={_al1}, zone={_zone[:20]}")
        return (_name, tkr, _ccy, _cur, _lo, _ath,
                _ttm, _fpe, _de, _mg, _note, STK, None)

    except Exception as _ex:
        print(f"[--add {tkr}] Failed: {_ex}")
        return None


if ADD_TICKERS and LIVE_REFRESH:
    for _add_ticker in ADD_TICKERS:
        _new_row = _fetch_new_ticker_row(_add_ticker)
        if _new_row:
            rows_data.append(_new_row)
elif ADD_TICKERS and not LIVE_REFRESH:
    print("[--add] LIVE_REFRESH=False — cannot auto-fetch new tickers. "
          "Set LIVE_REFRESH=True or add the ticker manually.")

# ── Write rows ─────────────────────────────────────────────────────────────────
row = 4
for rec in rows_data:
    (asset, ticker, ccy, current, macro_lo, macro_hi,
     ttm_pe, fwd_pe, de, margin, notes, rtype, manual) = rec

    # Skip any ticker the user excluded via --exclude
    if ticker in EXCLUDED_TICKERS:
        continue

    # Section row
    if rtype == S:
        ws.merge_cells(f"A{row}:AB{row}")
        c = ws.cell(row=row, column=1, value=asset)
        c.font = fnt(bold=True, color="FFFFFF", size=8)
        c.fill = fill(SECT)
        c.alignment = aln("left")
        c.border = bdr()
        ws.row_dimensions[row].height = 63 if asset.startswith("MARKET SENTIMENT") else 15
        row += 1
        continue

    if rtype == ISA_T:    bg = ISA_BG
    elif rtype == SIPP_T: bg = SIPP_BG
    elif rtype == MAN:    bg = MANUAL
    elif row % 2 == 0:   bg = ALT
    else:                bg = WHITE

    # ── Financial red-flag pre-computation (needed before col A is written) ──
    # Rules: D/E>1.0=HIGH LEVERAGE; margin<0=LOSS-MAKING; 0-5%=THIN MARGIN;
    #        TTM P/E>100=ELEVATED; Fwd P/E>80=STRETCHED
    # MAN rows that are real companies (not pre-IPO, not crypto, not macro indices) should also be assessed
    _MACRO_TICKERS = {"TBC", "XRP-USD", "ETH-USD", "BTC", "CRCL", "VIX", "TNX", "CL=F", "DXY"}
    _is_equity = (
        rtype in [STK, ISA_T, SIPP_T]
        or (rtype == MAN and ticker not in _MACRO_TICKERS)
    ) and ticker not in _MACRO_TICKERS
    fin_flags = []
    if _is_equity:
        if de is not None:
            if de > 2.0:
                fin_flags.append(f"EXTREME LEVERAGE D/E={de:.2f}")
            elif de > 1.0:
                fin_flags.append(f"HIGH LEVERAGE D/E={de:.2f}")
        if margin is not None:
            if margin < 0:
                fin_flags.append("LOSS-MAKING")
            elif margin < 0.05:
                fin_flags.append(f"THIN MARGIN {margin:.0%}")
        if ttm_pe is not None and ttm_pe > 100:
            fin_flags.append(f"HIGH TTM P/E {ttm_pe:.0f}x")
        if fwd_pe is not None and fwd_pe > 80:
            fin_flags.append(f"FWD P/E STRETCHED {fwd_pe:.0f}x")

    def cell(col, val, bold=False, color="000000", num_fmt=None, bg_ov=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = fnt(bold=bold, color=color, size=8)
        c.fill = fill(bg_ov or bg)
        c.alignment = aln("center" if col != 1 and col != 26 else "left")
        c.border = bdr()
        if num_fmt:
            c.number_format = num_fmt
        return c

    # A – Asset  (prefix "! " in bold red when financial rules are broken)
    if fin_flags:
        asset_label = f"! {asset}"
        asset_color = RED_C
        asset_bold  = True
    elif rtype == ISA_T:
        asset_label = asset
        asset_color = BLUE
        asset_bold  = True
    elif rtype == SIPP_T:
        asset_label = asset
        asset_color = "000000"
        asset_bold  = True
    else:
        asset_label = asset
        asset_color = "000000"
        asset_bold  = False
    cell(1, asset_label, bold=asset_bold, color=asset_color)
    # B – Ticker
    cell(2, ticker, bold=True)
    # C – CCY
    cell(3, ccy)
    # D – Current
    cell(4, current, bold=True, color="0000FF",
         num_fmt='#,##0.00' if current else None)

    # E – 2022 Bear Low
    c = ws.cell(row=row, column=5, value=macro_lo)
    c.font = fnt(color="0000FF", size=8)
    c.fill = fill(FIB_BG)
    c.alignment = aln()
    c.border = bdr()
    if macro_lo:
        c.number_format = '#,##0.00'

    # F – Cycle ATH
    c = ws.cell(row=row, column=6, value=macro_hi)
    c.font = fnt(color="0000FF", size=8)
    c.fill = fill(FIB_BG)
    c.alignment = aln()
    c.border = bdr()
    if macro_hi:
        c.number_format = '#,##0.00'

    # G – Fib Range
    if macro_lo and macro_hi:
        c = ws.cell(row=row, column=7, value=f"=F{row}-E{row}")
        c.number_format = '#,##0.00'
    else:
        c = ws.cell(row=row, column=7, value="—")
    c.font = fnt(size=8)
    c.fill = fill(FIB_BG)
    c.alignment = aln()
    c.border = bdr()

    # H/I/J/K – AL levels
    al_config = [
        (8,  0.382, AL1_BG, GRN),
        (9,  0.500, AL2_BG, AMB),
        (10, 0.618, AL3_BG, RED_C),
        (11, 0.786, D4_BG,  "880000"),
    ]
    for col, pct, al_bg, al_color in al_config:
        if rtype == MAN and manual:
            if col == 8:    val = manual[0]
            elif col == 9:  val = manual[1]
            elif col == 10: val = manual[2]
            else:           val = "N/A"
            c = ws.cell(row=row, column=col, value=val)
            c.font = fnt(bold=True, color=al_color, size=8, italic=True)
            c.fill = fill(al_bg)
            c.alignment = aln()
            c.border = bdr()
            if isinstance(val, (int, float)) and val:
                c.number_format = '#,##0.00'
        elif macro_lo and macro_hi:
            c = ws.cell(row=row, column=col, value=f"=F{row}-(G{row}*{pct})")
            c.font = fnt(color=al_color, size=8)
            c.fill = fill(al_bg)
            c.alignment = aln()
            c.border = bdr()
            c.number_format = '#,##0.00'
        else:
            c = ws.cell(row=row, column=col, value="—")
            c.font = fnt(size=8)
            c.fill = fill(al_bg)
            c.alignment = aln()
            c.border = bdr()

    # L – Upside to ATH
    if current and macro_hi:
        c = ws.cell(row=row, column=12, value=f"=(F{row}-D{row})/D{row}")
        c.number_format = '0.0%'
        c.font = fnt(bold=True, size=8)
    else:
        c = ws.cell(row=row, column=12, value="—")
        c.font = fnt(size=8)
    c.fill = fill("E8F5E9")
    c.alignment = aln()
    c.border = bdr()

    # M – Status vs macro Fib  (fin_flags already computed above; appended here)
    if rtype == MAN:
        status = "Manual levels\n(no macro Fib)"
        st_color = AMB
    elif current and macro_lo and macro_hi:
        al1 = fib(macro_hi, macro_lo, 0.382)
        if current >= al1:
            status = "Above AL1\nNo action yet"
            st_color = "555555"
        else:
            al2 = fib(macro_hi, macro_lo, 0.500)
            if current >= al2:
                status = "AL1 HIT\nACCUMULATE NOW"
                st_color = GRN
            else:
                al3 = fib(macro_hi, macro_lo, 0.618)
                if current >= al3:
                    status = "AL2 HIT\nSTRONG BUY NOW"
                    st_color = AMB
                else:
                    d4 = fib(macro_hi, macro_lo, 0.786)
                    if current >= d4:
                        status = "AL3 HIT\nBACK UP TRUCK"
                        st_color = RED_C
                    else:
                        status = "BELOW 78.6%\nDEEPEST VALUE"
                        st_color = "880000"
    else:
        status = "—"
        st_color = "555555"

    # Append financial flags to status text; escalate colour if normally neutral
    if fin_flags:
        status = status + "\n" + " | ".join(fin_flags)
        if st_color == "555555":    # was "No action yet" — upgrade to amber
            st_color = AMB

    c = ws.cell(row=row, column=13, value=status)
    c.font = fnt(bold=("HIT" in status or "BELOW" in status or bool(fin_flags)),
                  color=st_color, size=7.5)
    c.fill = fill(bg)
    c.alignment = aln()
    c.border = bdr()

    # N – TTM P/E  (red >100, amber 50-100, green <15)
    if ttm_pe is not None:
        pe_str = f"{ttm_pe:.0f}x"
        if ttm_pe > 100:
            pe_bg, pe_fc = "FFCDD2", RED_C
        elif ttm_pe > 50:
            pe_bg, pe_fc = AL2_BG, AMB
        elif ttm_pe < 15:
            pe_bg, pe_fc = AL1_BG, GRN
        else:
            pe_bg, pe_fc = FUND_BG, "000000"
    else:
        pe_str = "ETF" if rtype == ETF_T else "N/A"
        pe_bg, pe_fc = FUND_BG, "888888"
    c = ws.cell(row=row, column=14, value=pe_str)
    c.font = fnt(size=8, bold=(ttm_pe is not None and ttm_pe > 100), color=pe_fc)
    c.fill = fill(pe_bg)
    c.alignment = aln()
    c.border = bdr()

    # O – TTM EPS (col 15) / Q – Fwd EPS (col 17)
    # TTM EPS source priority:
    #   1. Alpha Vantage COMPANY_OVERVIEW (DilutedEPSTTM) — real, independent
    #   2. Derived: Price ÷ TTM P/E — displayed with ~ prefix, cannot verify P/E
    # Fwd EPS is always derived from Price ÷ Fwd P/E (analyst estimates, no independent source)
    _eps_row   = eps_data.get(ticker)
    _ttm_eps   = None
    _ttm_derived = True
    if _eps_row:
        _ttm_eps, _ttm_derived = _eps_row   # AV data: is_derived=False
    if _ttm_eps is None and ttm_pe and current:
        _ttm_eps     = round(current / ttm_pe, 2)
        _ttm_derived = True

    # Fwd EPS — always derived
    _fwd_eps = round(current / fwd_pe, 2) if (fwd_pe and current) else None

    # O – TTM EPS cell (col 15)
    if _ttm_eps is not None:
        _pfx       = "~$" if _ttm_derived else "$"   # ~ = derived, $ = real AV data
        _eps_label = f"{_pfx}{abs(_ttm_eps):.2f}" if _ttm_eps >= 0 else f"-{_pfx[1:]}{abs(_ttm_eps):.2f}"
        _eps_bg    = "FFCDD2" if _ttm_eps < 0 else FUND_BG
        _eps_fc    = RED_C    if _ttm_eps < 0 else ("555555" if _ttm_derived else "000000")
    else:
        _eps_label = "ETF" if rtype == ETF_T else "N/A"
        _eps_bg, _eps_fc = FUND_BG, "888888"
    c = ws.cell(row=row, column=15, value=_eps_label)
    c.font = fnt(size=8, color=_eps_fc, italic=_ttm_derived and _ttm_eps is not None)
    c.fill = fill(_eps_bg)
    c.alignment = aln()
    c.border = bdr()

    # P – Fwd P/E  (red >80, amber 40-80, green <20)
    if fwd_pe is not None:
        fpe_str = f"{fwd_pe:.0f}x"
        if fwd_pe > 80:
            fpe_bg, fpe_fc = "FFCDD2", RED_C
        elif fwd_pe > 40:
            fpe_bg, fpe_fc = AL2_BG, AMB
        elif fwd_pe < 20:
            fpe_bg, fpe_fc = AL1_BG, GRN
        else:
            fpe_bg, fpe_fc = FUND_BG, "000000"
    else:
        fpe_str = "ETF" if rtype == ETF_T else "N/A"
        fpe_bg, fpe_fc = FUND_BG, "888888"
    c = ws.cell(row=row, column=16, value=fpe_str)
    c.font = fnt(size=8, bold=(fwd_pe is not None and fwd_pe > 80), color=fpe_fc)
    c.fill = fill(fpe_bg)
    c.alignment = aln()
    c.border = bdr()

    # Q – Fwd EPS (col 17) — always derived from Price ÷ Fwd P/E
    if _fwd_eps is not None:
        _feps_label = f"~${abs(_fwd_eps):.2f}" if _fwd_eps >= 0 else f"-~${abs(_fwd_eps):.2f}"
        _feps_bg    = "FFCDD2" if _fwd_eps < 0 else FUND_BG
        _feps_fc    = RED_C    if _fwd_eps < 0 else "555555"
    else:
        _feps_label = "ETF" if rtype == ETF_T else "N/A"
        _feps_bg, _feps_fc = FUND_BG, "888888"
    c = ws.cell(row=row, column=17, value=_feps_label)
    c.font = fnt(size=8, color=_feps_fc, italic=(_fwd_eps is not None))
    c.fill = fill(_feps_bg)
    c.alignment = aln()
    c.border = bdr()

    # R – D/E  (red >1.0, amber 0.75-1.0, green <0.50)
    if de is not None:
        de_str = f"{de:.2f}"
        if de > 2.0:
            de_bg, de_fc = "FFCDD2", RED_C
        elif de > 1.0:
            de_bg, de_fc = "FFCDD2", RED_C
        elif de > 0.75:
            de_bg, de_fc = AL2_BG, AMB
        elif de < 0.50:
            de_bg, de_fc = AL1_BG, GRN
        else:
            de_bg, de_fc = FUND_BG, "000000"
    else:
        de_str = "ETF" if rtype == ETF_T else "N/A"
        de_bg, de_fc = FUND_BG, "888888"
    c = ws.cell(row=row, column=18, value=de_str)
    c.font = fnt(size=8, bold=(de is not None and de > 1.0), color=de_fc)
    c.fill = fill(de_bg)
    c.alignment = aln()
    c.border = bdr()

    # S – Net Margin  (red <0, amber 0-5%, green >20%)
    if margin is not None:
        if margin < 0:
            m_bg, m_fc, m_bold = "FFCDD2", RED_C, True
        elif margin < 0.05:
            m_bg, m_fc, m_bold = AL2_BG, AMB, False
        elif margin > 0.25:
            m_bg, m_fc, m_bold = AL1_BG, GRN, True
        elif margin > 0.10:
            m_bg, m_fc, m_bold = "E8F5E9", GRN, False
        else:
            m_bg, m_fc, m_bold = FUND_BG, "000000", False
        c = ws.cell(row=row, column=19, value=margin)
        c.number_format = '0%'
        c.font = fnt(size=8, bold=m_bold, color=m_fc)
        c.fill = fill(m_bg)
    else:
        c = ws.cell(row=row, column=19, value="ETF" if rtype == ETF_T else "N/A")
        c.font = fnt(size=8, color="888888")
        c.fill = fill(FUND_BG)
    c.alignment = aln()
    c.border = bdr()

    # ── MA columns T–X (20–24) + MA Trend Y (25) ─────────────────────────────
    w20_v, w50_v, m20_v, m50_v, d200_v = ma_data.get(ticker, (None, None, None, None, None))

    for col, ema_val in [(20, w20_v), (21, w50_v), (22, m20_v), (23, m50_v)]:
        if ema_val is not None:
            cell_bg, cell_fc = ma_cell_color(current, ema_val)
            c = ws.cell(row=row, column=col, value=ema_val)
            c.font = fnt(size=8, color=cell_fc)
            c.fill = fill(cell_bg)
            c.number_format = ma_num_fmt(ema_val)
        else:
            c = ws.cell(row=row, column=col, value="N/A")
            c.font = fnt(size=8, color="888888")
            c.fill = fill(MA_BG)
        c.alignment = aln()
        c.border = bdr()

    # X – D200 SMA (col 24)
    if d200_v is not None:
        cell_bg, cell_fc = ma_cell_color(current, d200_v)
        c = ws.cell(row=row, column=24, value=d200_v)
        c.font = fnt(size=8, color=cell_fc)
        c.fill = fill(cell_bg)
        c.number_format = ma_num_fmt(d200_v)
    else:
        c = ws.cell(row=row, column=24, value="N/A")
        c.font = fnt(size=8, color="888888")
        c.fill = fill(MA_BG)
    c.alignment = aln()
    c.border = bdr()

    # Y – MA Trend summary (col 25)
    trend = ma_trend_text(current, w20_v, w50_v, m20_v, m50_v, d200_v)
    c = ws.cell(row=row, column=25, value=trend)
    c.font = fnt(size=7.5)
    c.fill = fill(MA_BG)
    c.alignment = aln()
    c.border = bdr()

    # Z – Notes (col 26)
    c = ws.cell(row=row, column=26, value=notes)
    c.font = fnt(size=7.5, italic=(rtype == ISA_T))
    c.fill = fill(bg)
    c.alignment = aln("left")
    c.border = bdr()

    # AA – Analyst Consensus (col 27)  AB – Analyst Price Target (col 28)
    # Only for equities with Alpha Vantage data; N/A for ETFs, crypto, UK stocks, pre-IPO
    _an = analyst_data.get(ticker)
    if _an and rtype not in [ETF_T]:
        _cons, _n_ana, _pt = _an
        # Consensus cell
        _cons_label = f"{_cons} ({_n_ana})"
        if _cons == "Strong Buy":
            _cons_bg, _cons_fc = AL1_BG, GRN
        elif _cons == "Buy":
            _cons_bg, _cons_fc = "E8F5E9", "1B5E20"
        elif _cons == "Hold":
            _cons_bg, _cons_fc = AL2_BG, AMB
        else:
            _cons_bg, _cons_fc = "FFCDD2", RED_C
        c = ws.cell(row=row, column=27, value=_cons_label)
        c.font = fnt(size=8, bold=(_cons in ["Strong Buy", "Buy"]), color=_cons_fc)
        c.fill = fill(_cons_bg)
        c.alignment = aln()
        c.border = bdr()
        # Price target cell
        if current and current > 0:
            _upside = (_pt - current) / current
            _pt_label = f"${_pt:,.0f} ({_upside:+.0%})"
            if _upside > 0.30:
                _pt_bg, _pt_fc, _pt_bold = AL1_BG, GRN, True
            elif _upside > 0.10:
                _pt_bg, _pt_fc, _pt_bold = "E8F5E9", "1B5E20", False
            elif _upside >= 0:
                _pt_bg, _pt_fc, _pt_bold = AL2_BG, AMB, False
            else:
                _pt_bg, _pt_fc, _pt_bold = "FFCDD2", RED_C, False
        else:
            _pt_label, _pt_bg, _pt_fc, _pt_bold = "N/A", "F5F5F5", "888888", False
        c = ws.cell(row=row, column=28, value=_pt_label)
        c.font = fnt(size=8, bold=_pt_bold, color=_pt_fc)
        c.fill = fill(_pt_bg)
        c.alignment = aln()
        c.border = bdr()
    else:
        # ETF, crypto without AV data, UK stocks, pre-IPO
        _na_label = "ETF" if rtype == ETF_T else "N/A"
        for _col in [27, 28]:
            c = ws.cell(row=row, column=_col, value=_na_label)
            c.font = fnt(size=8, color="888888")
            c.fill = fill("F5F5F5")
            c.alignment = aln()
            c.border = bdr()

    ws.row_dimensions[row].height = 32
    row += 1

# Footer
ws.merge_cells(f"A{row}:AB{row}")
ws[f"A{row}"] = (
    f"Sources: Yahoo Finance (yfinance) — prices + MAs auto-refreshed {_run_date}  |  Analyst consensus: Alpha Vantage (manual refresh)  |  "
    "MACRO FIB: 2022 bear market monthly Low → cycle ATH monthly High  |  "
    "D200 SMA: simple 200-day rolling mean on daily close  |  "
    "W20/W50 EMA: daily resampled to weekly (Fri), EWM  |  "
    "M20/M50 EMA: 10Y monthly close, EWM  |  "
    "Green=price above MA (bullish), Red=price below MA (bearish)  |  "
    "GENIUS Act (Jul 2025): stablecoin yield prohibited — OCC rules Mar 2026  |  "
    "Not financial advice."
)
ws[f"A{row}"].font = fnt(italic=True, color="888888", size=7)
ws[f"A{row}"].fill = fill("F0F0F0")
ws[f"A{row}"].alignment = aln()
ws.row_dimensions[row].height = 12

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — Macro Fib Methodology
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Fib Methodology")

def s2(r, cols, bold=False, bg=None, colors_list=None):
    for ci, val in enumerate(cols, 1):
        c = ws2.cell(row=r, column=ci, value=val)
        col_color = colors_list[ci-1] if colors_list and ci <= len(colors_list) else "000000"
        c.font = fnt(bold=bold, color=col_color, size=8.5)
        c.fill = fill(bg or WHITE)
        c.alignment = aln("left")
        c.border = bdr()
    ws2.row_dimensions[r].height = 18

for ci, w in enumerate([22, 14, 14, 14, 13, 55], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

ws2.merge_cells("A1:F1")
ws2["A1"] = "Macro Fibonacci Retracement Framework — v20 Anchor Methodology"
ws2["A1"].font = fnt(bold=True, color="FFFFFF", size=12)
ws2["A1"].fill = fill(DARK)
ws2["A1"].alignment = aln()
ws2.row_dimensions[1].height = 22

r = 2
s2(r, ["Level", "Fib %", "Formula", "Example (META)", "Trigger Meaning", "Interpretation"],
   bold=True, bg="37474F", colors_list=["FFFFFF"]*6)
r += 1
# META macro: lo=87.40, hi=794.38
lvls = [
    ("AL1 — Accumulate",    "38.2%", "ATH−(Range×0.382)", f"${fib(794.38,87.40,0.382):,.2f}",
     "First macro pullback", "Start building. Institutional first stop. ~1/3 position.", AL1_BG, GRN),
    ("AL2 — Strong Buy",    "50.0%", "ATH−(Range×0.500)", f"${fib(794.38,87.40,0.500):,.2f}",
     "Macro midpoint",       "Increase size. 50% retrace = major support. ~1/3 more.", AL2_BG, AMB),
    ("AL3 — Back Up Truck", "61.8%", "ATH−(Range×0.618)", f"${fib(794.38,87.40,0.618):,.2f}",
     "Golden ratio retrace", "Maximum conviction. Final third. Historically rare at macro scale.", AL3_BG, RED_C),
    ("Deep Value",          "78.6%", "ATH−(Range×0.786)", f"${fib(794.38,87.40,0.786):,.2f}",
     "Near full retrace",    "Extreme caution OR maximum size if thesis intact. Structural breakdown risk.", D4_BG, "880000"),
]
for label, pct, formula, example, trigger, interp, bg_c, fc in lvls:
    s2(r, [label, pct, formula, example, trigger, interp], bg=bg_c,
       colors_list=[fc, fc, "000000", "000000", "000000", "000000"])
    r += 1

r += 1
s2(r, ["MACRO SWING ANCHORS (15 Apr 2026)", "", "", "", "", ""], bold=True, bg="F5F5F5")
r += 1
s2(r, ["Ticker", "2022 Bear Low", "Cycle ATH", "Current", "Macro AL1", "Macro Status"],
   bold=True, bg="37474F", colors_list=["FFFFFF"]*6)
r += 1

macro_swing_data = [
    ("MSTR",    13.26,    543.00,   132.36, f"${fib(543.00,13.26,0.382):,.2f}",      "BELOW 78.6% ($126.62) — deepest value. GTC BUY @ $114.", D4_BG, "880000"),
    ("COIN",    31.83,    444.65,   174.53, f"${fib(444.65,31.83,0.382):,.2f}",      "BELOW AL3 ($189.53) — back up truck. Exchange: GENIUS Act positive.", AL3_BG, RED_C),
    ("CEG",     40.73,    411.68,   291.72, f"${fib(411.68,40.73,0.382):,.2f}",      "Above macro AL1 ($270.53) — wait for pullback.", ALT, "555555"),
    ("TLN",     49.50,    451.28,   326.08, f"${fib(451.28,49.50,0.382):,.2f}",      "Above AL1 ($297.80). Amazon nuclear PPA. EXTREME LEVERAGE. Set alert at AL1.", AL1_BG, GRN),
    ("GEV",    118.56,    948.38,   991.12, f"${fib(948.38,118.56,0.382):,.2f}",     "Near ATH — above AL1 ($631.39). GE power/grid spin-off. Set alert at AL1.", ALT, "555555"),
    ("XOM",     52.75,    171.23,   152.64, f"${fib(171.23,52.75,0.382):,.2f}",      "Near ATH — well above macro AL1 ($117.00). Wait.", ALT, "555555"),
    ("COP",     61.71,    134.87,   123.62, f"${fib(134.87,61.71,0.382):,.2f}",      "AT ATH — wait for $107 area.", ALT, "555555"),
    ("MRVL",    34.65,    126.99,   131.30, f"${fib(126.99,34.65,0.382):,.2f}",      "Above macro AL1 ($91.71) — watch for AL1 test.", ALT, "555555"),
    ("PLTR",     5.92,    207.52,   132.37, f"${fib(207.52,5.92,0.382):,.2f}",       "Above macro AL1 ($130.51). Set alert at AL1.", ALT, "555555"),
    ("CCJ",     17.86,    135.24,   116.70, f"${fib(135.24,17.86,0.382):,.2f}",      "Above macro AL1 ($90.42). Hold — alert at $90.", ALT, "555555"),
    ("NKE",     77.22,    166.63,    44.20, f"${fib(166.63,77.22,0.382):,.2f}",      "BELOW 78.6% ($96.35) — deepest value. Price BELOW 2022 bear low ($77.22). CEO turnaround. Only buy on $77.22 reclaim.", D4_BG, "880000"),
    ("PANW",    68.37,    223.61,   162.51, f"${fib(223.61,68.37,0.382):,.2f}",      "AL1 HIT ($164.31) — accumulate. Between AL1 and AL2 ($146).", AL1_BG, GRN),
    ("RKLB",     3.48,     99.58,    70.62, f"${fib(99.58,3.48,0.382):,.2f}",        "Just above macro AL1 ($62.87). AL1 entry zone imminent.", AL1_BG, GRN),
    ("NVDA",    10.80,    212.17,   189.31, f"${fib(212.17,10.80,0.382):,.2f}",      "Above macro AL1 ($135.24). ISA hold. NVDA D200=$91.67 (rising, not breakdown).", ALT, "555555"),
    ("AMZN",    81.69,    258.60,   239.89, f"${fib(258.60,81.69,0.382):,.2f}",      "Above macro AL1 ($191.07). ISA hold. Watch $191.", ALT, "555555"),
    ("MSFT",   207.39,    552.24,   384.37, f"${fib(552.24,207.39,0.382):,.2f}",     "AL2 HIT ($379.82) — strong buy. Current between AL2 and AL3 ($339).", AL2_BG, AMB),
    ("META",    87.40,    794.38,   634.53, f"${fib(794.38,87.40,0.382):,.2f}",      "Above macro AL1 ($524.31) = 61.8% on TradingView. Monthly support. Set alert at $524.", AL1_BG, GRN),
    ("GOOGL",   82.66,    348.75,   321.31, f"${fib(348.75,82.66,0.382):,.2f}",      "Above macro AL1 ($246.04). Set alert at $246.", ALT, "555555"),
    ("NFLX",    16.27,    134.12,   103.16, f"${fib(134.12,16.27,0.382):,.2f}",      "Just above macro AL1 ($89.04). Very close to accumulate zone.", AL1_BG, GRN),
    ("XRP-USD",  0.29,      3.84,    1.377, f"${fib(3.84,0.29,0.382):,.4f}",         "BELOW AL3 ($1.64) — deep value. Macro ATH $3.84 = Jan 2018.", AL3_BG, RED_C),
    ("ETH-USD", 896.11,  4953.73,  2369.90,  f"${fib(4953.73,896.11,0.382):,.2f}",   "BELOW AL3 ($2,446) — deep value. Settlement layer for stablecoins.", AL3_BG, RED_C),
    ("BTC",   15599.05, 126198.07, 74442.23,  f"${fib(126198.07,15599.05,0.382):,.0f}", "ABOVE AL2 ($70,899) RECLAIMED. D200=$91,315 — still below D200. ISA rotation RE-ENGAGED.", AL2_BG, AMB),
    ("VWRP.L",  75.24,    140.92,   130.00, f"£{fib(140.92,75.24,0.382):,.2f}",     "Above macro AL1 (£115.30). D200=£122.60 — very close to D200 support.", ALT, "555555"),
    ("SMGB.L",  13.79,     55.76,    58.35, f"£{fib(55.76,13.79,0.382):,.2f}",      "Above macro AL1 (£39.69) — near ATH. 10% SIPP core holding.", ALT, "555555"),
    ("VUSA.L",  53.80,    107.01,    96.51, f"£{fib(107.01,53.80,0.382):,.2f}",      "ABOVE AL1 (£86.68) — hold/watch. S&P 500 UCITS ETF. Add at AL1 (£86.68) or AL2 (£80.41).", ALT, "555555"),
    ("SMT.L",  670.47,   1568.18,  1373.16,  f"{fib(1568.18,670.47,0.382):,.0f}p",   "AL1 HIT (1225p) — accumulate. Pre-IPO SpaceX/Anthropic vehicle.", AL1_BG, GRN),
    ("TSLA",   101.81,    498.83,   352.42, f"${fib(498.83,101.81,0.382):,.2f}",
     "AL1 ZONE ($347.17) — cautious accumulate. Robotaxi/Optimus thesis. Revenue -3.1% YoY. Fwd P/E 129x.", AL1_BG, GRN),
    ("INTC",    23.40,     68.49,    65.18, f"${fib(68.49,23.40,0.382):,.2f}",
     "ABOVE AL1 ($51.27) — approaching ATH. Turnaround play (IFS + CHIPS Act). Wait for AL1 ($51.27) pullback.", ALT, "555555"),
    ("MNTN.L",   0.83,      3.08,    1.92,  f"${fib(3.08,0.83,0.382):,.2f}",        "BELOW AL3 ($1.69) — deep value. Pure-play SpaceX/Anthropic.", AL3_BG, RED_C),
]
for row_data in macro_swing_data:
    tick, lo, hi, cur, al1_val, status, bg_c, fc = row_data
    lo_str  = f"{lo:,.2f}"  if lo  else "N/A"
    hi_str  = f"{hi:,.2f}"  if hi  else "N/A"
    cur_str = f"{cur:,.2f}" if cur else "N/A"
    s2(r, [tick, lo_str, hi_str, cur_str, al1_val, status],
       bg=bg_c, colors_list=["000000", "0000FF", "0000FF", "000000", fc, fc])
    r += 1

r += 1
s2(r, ["v17 CHANGE: All prices refreshed 14 Apr 2026. BTC ABOVE AL2 ($70,899) — ISA rotation thesis RE-ENGAGED. VIX 18.6 — normalising.", "", "", "", "", ""],
   bg="E3F2FD")
r += 1
s2(r, ["v18 CHANGE: VUSA.L added as SIPP play. Vanguard S&P 500 UCITS ETF. Fib: Jun 2022 low £53.80 → ATH £107.01. AL1=£86.68 | AL2=£80.41 | AL3=£74.13. W20=£94.95 / W50=£92.41 / M20=£89.19 / M50=£77.96 / D200=£94.06.", "", "", "", "", ""],
   bg="E3F2FD")
r += 1
s2(r, ["v19 CHANGE: INTC (Intel Corporation) added as ISA play. Macro Fib: Oct 2022 low $23.40 → Feb 2021 ATH $68.49. AL1=$51.27 | AL2=$45.94 | AL3=$40.62 | 78.6%=$33.05. W20=$46.06 / W50=$36.82 / M20=$36.03 / M50=$31.91 / D200=$36.41. Loss-making (net margin -0.5%). Wait for AL1 ($51.27) pullback before entry.", "", "", "", "", ""],
   bg="E3F2FD")
r += 1
s2(r, ["v20 CHANGE: NKE (Nike, Inc.) added to Watchlist. Macro Fib: Oct 2022 low $77.22 → ATH $166.63 Nov 2021. AL1=$132.47 | AL2=$121.92 | AL3=$111.37 | 78.6%=$96.35. EXTREME: current $44.20 BELOW 2022 bear low ($77.22). W20=$56.16 / W50=$59.90 / M20=$66.92 / M50=$80.31 / D200=$65.17. Thin margin 4.8%. D/E=0.98. CEO Hill turnaround. Only accumulate on $77.22 reclaim.", "", "", "", "", ""],
   bg="E3F2FD")
r += 1
s2(r, ["TSLA Fib: 2022 low $101.81 → ATH $498.83. AL1=$347.17 | AL2=$300.32 | AL3=$253.47 | 78.6%=$186.77.", "", "", "", "", ""],
   bg="E3F2FD")
r += 1
s2(r, ["All data sourced from Yahoo Finance (daily 1Y + monthly 10Y) for consistency. Weekly EMAs from daily resampled to Friday close.", "", "", "", "", ""],
   bg="E3F2FD")


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 0 — Daily Summary  (inserted at front; becomes first tab)
# Shows all assets grouped by Fib zone: BUY ZONES → APPROACHING → WATCHING
# ─────────────────────────────────────────────────────────────────────────────
ws_s = wb.create_sheet("Daily Summary", 0)
ws_s.freeze_panes = "A4"
ws_s.sheet_view.zoomScale = 95

S_COLS = [
    "Asset", "Ticker", "CCY",
    "Current\nPrice",
    "Zone / Action",
    "AL1\n38.2%", "AL2\n50.0%", "AL3\n61.8%", "Deep Value\n78.6%",
    "Gap to\nAL1",
    "Upside\nto ATH",
    "Analyst\nConsensus",
    "Analyst\nPrice Target",
    "Key Thesis",
]
S_WIDTHS = [26, 8, 5, 11, 20, 12, 12, 12, 12, 10, 10, 14, 13, 52]

# Zone rank: 0=deepest, 1=AL3, 2=AL2, 3=AL1, 4=approaching, 5=above, 9=no fib
ZONE_STYLES = {
    # (bg, font_color, label)
    0: ("880000", "FFFFFF", "BELOW 78.6%  DEEPEST VALUE"),
    1: ("FFCDD2", "880000", "AL3 HIT  BACK UP TRUCK"),
    2: ("FFE0B2", "BF360C", "AL2 HIT  STRONG BUY"),
    3: ("E8F5E9", "1B5E20", "AL1 HIT  ACCUMULATE"),
    4: ("FFF9C4", "F57F17", "APPROACHING AL1"),
    5: ("F5F5F5", "757575", "Above AL1 — Watching"),
    9: ("ECEFF1", "607D8B", "No Fib / ETF"),
}

# Non-tradeable sentiment gauges — show zone only (no action verb) in the
# Daily Summary "Zone / Action" column because we cannot actually buy these.
SENTIMENT_TICKERS = {"CL=F", "VIX", "TNX", "TLT", "DXY"}

# Zone-only labels used when the ticker is in SENTIMENT_TICKERS.
# Colour coding is unchanged — only the displayed text differs.
ZONE_LABELS_SENTIMENT = {
    0: "BELOW 78.6%",
    1: "AL3 HIT",
    2: "AL2 HIT",
    3: "AL1 HIT",
    4: "APPROACHING AL1",
    5: "Above AL1",
    9: "No Fib / ETF",
}

def _first_sentence(txt):
    """Extract first meaningful clause from notes (up to first full stop or pipe)."""
    if not txt:
        return ""
    for sep in [". ", " — ", " | "]:
        idx = txt.find(sep)
        if idx > 0 and idx < 180:
            return txt[:idx + (1 if sep == ". " else 0)].strip()
    return txt[:150].strip()

# Build summary rows from rows_data
sum_rows = []
for rec in rows_data:
    (asset, ticker, ccy, current, macro_lo, macro_hi,
     ttm_pe, fwd_pe, de, margin, notes, rtype, manual) = rec

    # Skip any ticker the user excluded via --exclude
    if ticker in EXCLUDED_TICKERS:
        continue
    if rtype == S or current is None:
        continue

    # Fib levels
    if rtype == MAN and manual and manual[0] is not None:
        al1_v, al2_v, al3_v = manual
        d4_v = None
    elif macro_lo and macro_hi:
        al1_v = fib(macro_hi, macro_lo, 0.382)
        al2_v = fib(macro_hi, macro_lo, 0.500)
        al3_v = fib(macro_hi, macro_lo, 0.618)
        d4_v  = fib(macro_hi, macro_lo, 0.786)
    else:
        al1_v = al2_v = al3_v = d4_v = None

    # Determine zone
    if al1_v is None:
        zone_rank = 9
        gap_to_al1 = None
    else:
        gap_to_al1 = (current - al1_v) / al1_v   # negative = below AL1 (in zone)
        if d4_v and current < d4_v:
            zone_rank = 0
        elif al3_v and current < al3_v:
            zone_rank = 1
        elif al2_v and current < al2_v:
            zone_rank = 2
        elif current < al1_v:
            zone_rank = 3
        elif gap_to_al1 <= 0.15:
            zone_rank = 4
        else:
            zone_rank = 5

    upside = (macro_hi - current) / macro_hi if macro_hi else None
    an     = analyst_data.get(ticker)

    sum_rows.append({
        "asset": asset, "ticker": ticker, "ccy": ccy, "current": current,
        "zone_rank": zone_rank, "al1": al1_v, "al2": al2_v,
        "al3": al3_v, "d4": d4_v, "gap": gap_to_al1,
        "upside": upside, "analyst": an, "rtype": rtype,
        "note": _first_sentence(notes),
    })

# Sort: within buy zones (0-3) sort most extreme first (lowest gap);
# approaching (4) sort closest first; watching (5,9) alphabetical
sum_rows.sort(key=lambda x: (
    x["zone_rank"],
    x["gap"] if x["gap"] is not None else 99,
))

# ── Title & subtitle ──────────────────────────────────────────────────────────
ws_s.merge_cells("A1:N1")
ws_s["A1"] = "SIPP / ISA — Daily Action Summary  |  Fibonacci Alert Levels"
ws_s["A1"].font = fnt(bold=True, color="FFFFFF", size=13)
ws_s["A1"].fill = fill(DARK)
ws_s["A1"].alignment = aln()
ws_s.row_dimensions[1].height = 22

ws_s.merge_cells("A2:N2")
ws_s["A2"] = (
    f"10 Apr 2026  |  "
    "IN BUY ZONES: price has retraced to AL1/AL2/AL3 — action required  |  "
    "APPROACHING: within 15% of AL1 — set alerts  |  "
    "WATCHING: above AL1 — no action  |  "
    "Gap to AL1: negative = already in zone (price below AL1)  |  "
    "Upside = % from current to cycle ATH"
)
ws_s["A2"].font = fnt(italic=True, color="555555", size=7.5)
ws_s["A2"].fill = fill("F0F0F0")
ws_s["A2"].alignment = aln()
ws_s.row_dimensions[2].height = 13

# ── Column headers ────────────────────────────────────────────────────────────
for ci, (h, w) in enumerate(zip(S_COLS, S_WIDTHS), 1):
    c = ws_s.cell(row=3, column=ci, value=h)
    c.font  = fnt(bold=True, color="FFFFFF", size=7.5)
    c.fill  = fill(DARK)
    c.alignment = aln()
    c.border = bdr()
    ws_s.column_dimensions[get_column_letter(ci)].width = w
ws_s.row_dimensions[3].height = 30

# ── Section headers ───────────────────────────────────────────────────────────
SECTION_DEFS = [
    (range(0, 4), "IN BUY ZONES  —  Price has retraced to a Fib alert level — action required today",  "B71C1C"),
    (range(4, 5), "APPROACHING AL1  —  Within 15% of the 38.2% accumulate level — set price alerts",   "E65100"),
    (range(5, 6), "WATCHING  —  Above AL1 — no entry yet; monitor for pullback to alert levels",        "37474F"),
    (range(9,10), "NO FIB / ETF  —  Manual levels or ETF; refer to Alert Levels tab",                  "455A64"),
]

def _section_row(ws, row, text, bg):
    ws.merge_cells(f"A{row}:N{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(bold=True, color="FFFFFF", size=8)
    c.fill = fill(bg)
    c.alignment = aln("left")
    c.border = bdr()
    ws.row_dimensions[row].height = 14

# ── Write data rows ───────────────────────────────────────────────────────────
srow = 4
last_group = -1

for d in sum_rows:
    zr = d["zone_rank"]

    # Section header when group changes
    if zr != last_group:
        for grp_range, label, sbg in SECTION_DEFS:
            if zr in grp_range:
                _section_row(ws_s, srow, label, sbg)
                srow += 1
                break
        last_group = zr

    zbg, zfc, zlabel = ZONE_STYLES[zr]

    # Row background (alternating within zone)
    row_bg = zbg if zr <= 3 else ("FFFFFF" if srow % 2 == 0 else "F9F9F9")

    def sc(col, val, bold=False, color="000000", num_fmt=None, bg_ov=None, align="center"):
        c = ws_s.cell(row=srow, column=col, value=val)
        c.font = fnt(bold=bold, color=color, size=8)
        c.fill = fill(bg_ov or row_bg)
        c.alignment = aln(align)
        c.border = bdr()
        if num_fmt:
            c.number_format = num_fmt
        return c

    # A – Asset
    sc(1,  d["asset"],   bold=(zr <= 3), color=(zfc if zr <= 3 else "000000"), align="left")
    # B – Ticker
    sc(2,  d["ticker"],  bold=True, color=(zfc if zr <= 3 else "0000FF"))
    # C – CCY
    sc(3,  d["ccy"])
    # D – Current price
    sc(4,  d["current"], bold=True, color="0000FF", num_fmt='#,##0.00')

    # E – Zone label (colour-coded)
    # Non-tradeable sentiment gauges (VIX, TNX, DXY, TLT, CL=F) show zone only —
    # no action verb, because we cannot actually buy them. Colour coding unchanged.
    display_label = (
        ZONE_LABELS_SENTIMENT[zr] if d["ticker"] in SENTIMENT_TICKERS else zlabel
    )
    c = ws_s.cell(row=srow, column=5, value=display_label)
    c.font = fnt(bold=(zr <= 3), color=zfc, size=7.5)
    c.fill = fill(zbg)
    c.alignment = aln()
    c.border = bdr()

    # F/G/H/I – AL levels
    al_vals = [
        (6,  d["al1"], AL1_BG, GRN),
        (7,  d["al2"], AL2_BG, AMB),
        (8,  d["al3"], AL3_BG, RED_C),
        (9,  d["d4"],  D4_BG,  "880000"),
    ]
    for col, val, al_bg, al_fc in al_vals:
        if val:
            c = ws_s.cell(row=srow, column=col, value=val)
            c.font = fnt(size=8, bold=(zr <= 3), color=al_fc)
            c.fill = fill(al_bg)
            c.number_format = '#,##0.00'
        else:
            c = ws_s.cell(row=srow, column=col, value="—")
            c.font = fnt(size=8, color="888888")
            c.fill = fill(al_bg)
        c.alignment = aln()
        c.border = bdr()

    # J – Gap to AL1
    if d["gap"] is not None:
        gap_label = f"{d['gap']:+.1%}"
        if d["gap"] < 0:
            gap_bg, gap_fc = AL1_BG, GRN   # in zone — green
        elif d["gap"] < 0.08:
            gap_bg, gap_fc = AL2_BG, AMB   # very close — amber
        else:
            gap_bg, gap_fc = "F5F5F5", "757575"
        c = ws_s.cell(row=srow, column=10, value=gap_label)
        c.font = fnt(size=8, bold=(d["gap"] < 0), color=gap_fc)
        c.fill = fill(gap_bg)
    else:
        c = ws_s.cell(row=srow, column=10, value="—")
        c.font = fnt(size=8, color="888888")
        c.fill = fill("F5F5F5")
    c.alignment = aln()
    c.border = bdr()

    # K – Upside to ATH
    if d["upside"] is not None:
        up_str = f"{d['upside']:+.1%}"
        up_bg  = AL1_BG if d["upside"] > 0.30 else ("FFF8E1" if d["upside"] > 0.10 else "F5F5F5")
        up_fc  = GRN    if d["upside"] > 0.30 else (AMB      if d["upside"] > 0.10 else "757575")
        c = ws_s.cell(row=srow, column=11, value=up_str)
        c.font = fnt(size=8, color=up_fc)
        c.fill = fill(up_bg)
    else:
        c = ws_s.cell(row=srow, column=11, value="—")
        c.font = fnt(size=8, color="888888")
        c.fill = fill("F5F5F5")
    c.alignment = aln()
    c.border = bdr()

    # L – Analyst Consensus
    an = d["analyst"]
    if an and d["rtype"] not in [ETF_T]:
        cons, n_ana, pt = an
        cons_label = f"{cons} ({n_ana})"
        if cons == "Strong Buy":
            c_bg, c_fc = AL1_BG, GRN
        elif cons == "Buy":
            c_bg, c_fc = "E8F5E9", "1B5E20"
        elif cons == "Hold":
            c_bg, c_fc = AL2_BG, AMB
        else:
            c_bg, c_fc = "FFCDD2", RED_C
        c = ws_s.cell(row=srow, column=12, value=cons_label)
        c.font = fnt(size=8, bold=(cons in ["Strong Buy", "Buy"]), color=c_fc)
        c.fill = fill(c_bg)
    else:
        c = ws_s.cell(row=srow, column=12, value="ETF" if d["rtype"] == ETF_T else "N/A")
        c.font = fnt(size=8, color="888888")
        c.fill = fill("F5F5F5")
    c.alignment = aln()
    c.border = bdr()

    # M – Analyst Price Target
    if an and d["rtype"] not in [ETF_T] and d["current"]:
        cons, n_ana, pt = an
        upside_pt = (pt - d["current"]) / d["current"]
        pt_label  = f"${pt:,.0f} ({upside_pt:+.0%})"
        if upside_pt > 0.30:
            pt_bg, pt_fc = AL1_BG, GRN
        elif upside_pt > 0.10:
            pt_bg, pt_fc = "E8F5E9", "1B5E20"
        elif upside_pt >= 0:
            pt_bg, pt_fc = AL2_BG, AMB
        else:
            pt_bg, pt_fc = "FFCDD2", RED_C
        c = ws_s.cell(row=srow, column=13, value=pt_label)
        c.font = fnt(size=8, bold=(upside_pt > 0.30), color=pt_fc)
        c.fill = fill(pt_bg)
    else:
        c = ws_s.cell(row=srow, column=13, value="N/A")
        c.font = fnt(size=8, color="888888")
        c.fill = fill("F5F5F5")
    c.alignment = aln()
    c.border = bdr()

    # N – Key thesis (first sentence)
    c = ws_s.cell(row=srow, column=14, value=d["note"])
    c.font = fnt(size=7.5, italic=(zr >= 5))
    c.fill = fill(row_bg)
    c.alignment = aln("left")
    c.border = bdr()

    ws_s.row_dimensions[srow].height = 28
    srow += 1

# ── Footer ────────────────────────────────────────────────────────────────────
ws_s.merge_cells(f"A{srow}:N{srow}")
ws_s[f"A{srow}"] = (
    "Prices: Yahoo Finance MCP (10 Apr 2026)  |  "
    "Fib: 2022 bear market low → cycle ATH  |  "
    "AL1=38.2%  AL2=50%  AL3=61.8%  Deep Value=78.6%  |  "
    "Gap to AL1: negative = price already below AL1 (in zone)  |  "
    "Not financial advice."
)
ws_s[f"A{srow}"].font  = fnt(italic=True, color="888888", size=7)
ws_s[f"A{srow}"].fill  = fill("F0F0F0")
ws_s[f"A{srow}"].alignment = aln()
ws_s.row_dimensions[srow].height = 12

# ─────────────────────────────────────────────────────────────────────────────
# Sheet — How to Run
# ─────────────────────────────────────────────────────────────────────────────
ws_help = wb.create_sheet("How to Run")
ws_help.sheet_view.showGridLines = False
ws_help.column_dimensions["A"].width = 90

def _h(row, text, bold=False, size=11, color="000000", bg=None, indent=0):
    c = ws_help.cell(row=row, column=1, value=(" " * indent) + text)
    c.font = Font(name="Calibri", size=size, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    ws_help.row_dimensions[row].height = 18 if not bold else 22
    return row + 1

r = 1
r = _h(r, "How to Run This Spreadsheet", bold=True, size=14, color="1F4E79", bg="DEEAF1")
r = _h(r, "")
r = _h(r, "WHAT IT DOES", bold=True, size=11, color="2E75B6")
r = _h(r, "This spreadsheet is generated by a Python script. Each time you run it, a fresh timestamped "
          ".xlsx is created in the same folder. Current prices and all five moving averages "
          "(W20/W50 EMA, M20/M50 EMA, D200 SMA) are fetched live from Yahoo Finance on every run. "
          "Analyst consensus and price targets (Alpha Vantage) remain hardcoded — update manually "
          "or via Cowork on demand. 2022 bear lows, cycle ATHs, and thesis notes are also hardcoded.", size=10, indent=2)
r = _h(r, "")
r = _h(r, "REQUIREMENTS", bold=True, size=11, color="2E75B6")
r = _h(r, "Python 3  (pre-installed on macOS — type  python3 --version  in Terminal to confirm)", size=10, indent=2)
r = _h(r, "openpyxl  library  (one-time install — see below)", size=10, indent=2)
r = _h(r, "")
r = _h(r, "ONE-TIME SETUP  (only needed once)", bold=True, size=11, color="2E75B6")
r = _h(r, "Open Terminal  (Cmd + Space, type Terminal, press Enter)", size=10, indent=2)
r = _h(r, "Run:   pip3 install openpyxl yfinance pandas", size=10, indent=2, color="7030A0")
r = _h(r, "That's it. You never need to do this again.", size=10, indent=2)
r = _h(r, "")
r = _h(r, "RUNNING THE SCRIPT", bold=True, size=11, color="2E75B6")
r = _h(r, "Open Terminal", size=10, indent=2)
r = _h(r, "Run the command below (copy and paste it):", size=10, indent=2)
r = _h(r, "")
ws_help.row_dimensions[r].height = 22
c = ws_help.cell(row=r, column=1,
                  value='    python3 ~/Documents/My\\ Documents/Finance/Investments/build_alert_levels_v19.py  # v20 script')
c.font = Font(name="Courier New", size=11, bold=True, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F4E79")
c.alignment = Alignment(vertical="center")
ws_help.row_dimensions[r].height = 24
r += 1
r = _h(r, "")
r = _h(r, "The new spreadsheet will appear in the same Investments folder, named:", size=10, indent=2)
r = _h(r, "SIPP_Alert_Levels_v20_DDMMYY-HHMM.xlsx", size=10, indent=4, color="7030A0")
r = _h(r, "")
r = _h(r, "LIVE DATA REFRESH", bold=True, size=11, color="2E75B6")
r = _h(r, "Prices and MAs are fetched automatically every time you run the script (LIVE_REFRESH = True at the top).", size=10, indent=2)
r = _h(r, "To disable live fetch (e.g. no internet): open the script and set  LIVE_REFRESH = False  at the top.", size=10, indent=2)
r = _h(r, "What IS refreshed live (Yahoo Finance / yfinance):", size=10, indent=2, color="375623")
r = _h(r, "    Prices  |  D200 SMA  |  W20/W50 weekly EMA  |  M20/M50 monthly EMA", size=10, indent=4, color="375623")
r = _h(r, "    TTM P/E  |  Forward P/E  |  D/E ratio  |  Net margin", size=10, indent=4, color="375623")
r = _h(r, "    Analyst consensus (Strong Buy/Buy/Hold/Sell)  |  Analyst mean price target", size=10, indent=4, color="375623")
r = _h(r, "What is NOT refreshed (update manually or via Cowork):", size=10, indent=2, color="7030A0")
r = _h(r, "    2022 bear lows  |  Cycle ATHs  |  Thesis / notes", size=10, indent=4, color="7030A0")
r = _h(r, "")
r = _h(r, "CLI OPTIONS", bold=True, size=11, color="2E75B6")
r = _h(r, "The script supports two optional command-line flags:", size=10, indent=2)
r = _h(r, "  --exclude TICKER [TICKER ...]   Omit one or more tickers from this run (no data fetched, no rows written)", size=10, indent=2, color="7030A0")
r = _h(r, "  --add    TICKER [TICKER ...]   Fetch a ticker live and append it to the Watchlist section", size=10, indent=2, color="375623")
r = _h(r, "Examples:", size=10, indent=2)
r = _h(r, "  python3 build_alert_levels_v19.py --exclude NKE INTC", size=10, indent=4, color="1F4E79")
r = _h(r, "  python3 build_alert_levels_v19.py --add ORCL ADBE", size=10, indent=4, color="1F4E79")
r = _h(r, "  python3 build_alert_levels_v19.py --add ORCL --exclude INTC", size=10, indent=4, color="1F4E79")
r = _h(r, "--add requires LIVE_REFRESH=True (internet access). 2022 bear lows, cycle ATHs, and notes are auto-generated and can be refined via Cowork.", size=10, indent=2)
r = _h(r, "")
r = _h(r, "SCRIPT VERSIONS", bold=True, size=11, color="2E75B6")
r = _h(r, "Each new version is a self-contained Python file (v17, v18, v19, v20, ...). The latest version is always the one to use.", size=10, indent=2)
r = _h(r, "Previous versions are kept as a backup — do not delete them.", size=10, indent=2)
r = _h(r, "")
r = _h(r, "SCRIPT LOCATION", bold=True, size=11, color="2E75B6")
r = _h(r, "~/Documents/My Documents/Finance/Investments/build_alert_levels_v19.py  (v20)", size=10, indent=2, color="7030A0")
r = _h(r, "")
r = _h(r, "NOT FINANCIAL ADVICE — for personal tracking and research only.", bold=True, size=9, color="888888")

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
