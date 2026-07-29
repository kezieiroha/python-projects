"""End-to-end offline workbook contract tests.

These tests exercise the stable command entrypoint with deterministic runtime
flags and validate the regressions found during the redesign review.
"""

import contextlib
import io
import json
import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

import build_alert_levels


ROOT = Path(__file__).resolve().parents[1]
RUN_DATE = "2026-07-28T09:30"


class Phase0WorkbookContractTest(unittest.TestCase):
    def build_workbook(self, *extra_args):
        tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(tmpdir.cleanup)
        output_path = Path(tmpdir.name) / "workbook.xlsx"
        args = [
            "--offline",
            "--run-date", RUN_DATE,
            "--output", str(output_path),
            *extra_args,
        ]
        with contextlib.redirect_stdout(io.StringIO()):
            build_alert_levels.main(args)
        return output_path, load_workbook(output_path, data_only=False)

    def test_import_has_no_side_effect_execution(self):
        self.assertTrue(callable(build_alert_levels.main))

    def test_offline_workbook_contract_covers_closed_defects(self):
        _output_path, workbook = self.build_workbook()

        self.assertEqual(
            workbook.sheetnames,
            ["Daily Summary", "Alert Levels", "Fib Methodology", "How to Run"],
        )

        all_text = [
            cell.value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]

        self.assertEqual(
            sum(text.startswith("IN BUY ZONES") for text in all_text),
            1,
        )
        self.assertFalse(
            any(
                stale in text
                for text in all_text
                for stale in [
                    "10 Apr 2026",
                    "14 Apr 2026",
                    "15 Apr 2026",
                    "Yahoo Finance MCP",
                ]
            )
        )

        summary = workbook["Daily Summary"]
        alerts = workbook["Alert Levels"]
        self.assertIn("28 Jul 2026", summary["A2"].value)
        self.assertIn("Static JSON fallback data", alerts["A2"].value)
        self.assertEqual(alerts["A1"].fill.fgColor.rgb, "001A1A2E")
        for row in range(4, alerts.max_row + 1):
            for col in range(7, 13):
                value = alerts.cell(row=row, column=col).value
                self.assertFalse(
                    isinstance(value, str) and value.startswith("="),
                    f"Alert Levels {alerts.cell(row=row, column=col).coordinate} should be snapshot output",
                )

        sample = None
        for row in range(4, alerts.max_row + 1):
            current = alerts.cell(row=row, column=4).value
            macro_hi = alerts.cell(row=row, column=6).value
            upside_value = alerts.cell(row=row, column=12).value
            if current and macro_hi and isinstance(upside_value, (int, float)):
                self.assertEqual(alerts.cell(row=row, column=4).number_format, "#,##0.00")
                self.assertAlmostEqual(upside_value, (macro_hi - current) / current)
                self.assertEqual(alerts.cell(row=row, column=12).number_format, "0.0%")
                sample = (
                    alerts.cell(row=row, column=2).value,
                    f"{((macro_hi - current) / current):+.1%}",
                )
                break
        self.assertIsNotNone(sample)

        ticker, expected_upside = sample
        summary_upside = None
        for row in range(4, summary.max_row + 1):
            if summary.cell(row=row, column=2).value == ticker:
                summary_upside = summary.cell(row=row, column=11).value
                break
        self.assertEqual(summary_upside, expected_upside)

        eps_labels = [
            alerts.cell(row=row, column=15).value
            for row in range(4, alerts.max_row + 1)
            if isinstance(alerts.cell(row=row, column=15).value, str)
            and "$" in alerts.cell(row=row, column=15).value
        ]
        self.assertGreater(len(eps_labels), 0)
        self.assertFalse(
            any(label.startswith("$") or label.startswith("-$") for label in eps_labels)
        )

    def test_remove_added_ticker_affects_same_run_render(self):
        tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(tmpdir.cleanup)
        data_dir = Path(tmpdir.name) / "data"
        shutil.copytree(ROOT / "workbook" / "data", data_dir)

        output_path = Path(tmpdir.name) / "removed.xlsx"
        with contextlib.redirect_stdout(io.StringIO()):
            build_alert_levels.main([
                "--offline",
                "--run-date", RUN_DATE,
                "--output", str(output_path),
                "--data-dir", str(data_dir),
                "--remove", "SPCX",
            ])

        portfolio_rows = json.loads((data_dir / "portfolio_rows.json").read_text())
        self.assertFalse(
            any((row.get("ticker") or "").upper() == "SPCX" for row in portfolio_rows)
        )

        workbook = load_workbook(output_path, data_only=False)
        self.assertFalse(
            any(cell.value == "SPCX" for row in workbook["Daily Summary"].iter_rows() for cell in row)
        )
        self.assertFalse(
            any(cell.value == "SPCX" for row in workbook["Alert Levels"].iter_rows() for cell in row)
        )

    def test_same_ticker_added_rows_render_as_distinct_assets(self):
        tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(tmpdir.cleanup)
        data_dir = Path(tmpdir.name) / "data"
        shutil.copytree(ROOT / "workbook" / "data", data_dir)

        portfolio_path = data_dir / "portfolio_rows.json"
        rows = json.loads(portfolio_path.read_text())
        rows.extend([
            {
                "asset": "First",
                "ticker": "ABC",
                "ccy": "USD",
                "current": 10,
                "macro_lo": 5,
                "macro_hi": 20,
                "notes": "first",
                "rtype": "STOCK",
                "manual": None,
                "source": "added",
            },
            {
                "asset": "Second",
                "ticker": "ABC",
                "ccy": "USD",
                "current": 12,
                "macro_lo": 6,
                "macro_hi": 24,
                "notes": "second",
                "rtype": "STOCK",
                "manual": None,
                "source": "added",
            },
        ])
        portfolio_path.write_text(json.dumps(rows, indent=2))

        output_path = Path(tmpdir.name) / "same_ticker.xlsx"
        with contextlib.redirect_stdout(io.StringIO()):
            build_alert_levels.main([
                "--offline",
                "--run-date", RUN_DATE,
                "--output", str(output_path),
                "--data-dir", str(data_dir),
            ])

        workbook = load_workbook(output_path, data_only=False)
        rendered_assets = [
            cell.value
            for row in workbook["Alert Levels"].iter_rows()
            for cell in row
            if cell.value in {"First", "Second"}
        ]
        self.assertEqual(rendered_assets, ["First", "Second"])


if __name__ == "__main__":
    unittest.main()
