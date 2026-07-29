"""Live provider smoke check — the only thing here that touches the network.

Every test in this directory runs `--offline`, which means the suite cannot
detect a breaking yfinance release: the payload shapes it returns and the
`yf.download` signature are exactly what the offline tests never exercise. A
dependency bump can go green through CI and still leave every price blank.

This script closes that gap. It is deliberately NOT named `test_*.py`, so
neither `pytest` nor `unittest discover` collects it — it only runs when
invoked explicitly:

    python tests/live_smoke.py

Run it against a trimmed ticker list so a smoke check costs three provider
calls rather than sixty-seven.
"""

from __future__ import annotations

import json
import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import build_alert_levels  # noqa: E402

# Preferred smoke tickers: liquid, long-listed, unlikely to be delisted, so a
# failure means the provider integration broke rather than the ticker did.
PREFERRED = ["AAPL", "MSFT", "GOOGL", "AMZN", "META"]
SMOKE_COUNT = 3


def _pick_smoke_tickers(data_dir):
    """Choose tickers that are both fetched and rendered as portfolio rows.

    A ticker in `yahoo_fetch` but absent from the portfolio would be fetched and
    then have no row to assert against, which reads as a failure when nothing is
    wrong. Derived from the data rather than hardcoded so it survives portfolio
    edits.
    """
    fetched = set(json.loads((data_dir / "market_sources.json").read_text())
                  .get("yahoo_fetch", []))
    rows = json.loads((data_dir / "portfolio_rows.json").read_text())
    rendered = {r.get("ticker") for r in rows if r.get("ticker")}
    eligible = fetched & rendered

    picked = [t for t in PREFERRED if t in eligible]
    picked += sorted(eligible - set(picked))
    return picked[:SMOKE_COUNT]


def _trimmed_data_dir(tmp):
    """Copy workbook/data/ and cut the fetch list down to the smoke tickers."""
    data_dir = Path(tmp) / "data"
    shutil.copytree(ROOT / "workbook" / "data", data_dir)

    sources_path = data_dir / "market_sources.json"
    sources = json.loads(sources_path.read_text())
    smoke = _pick_smoke_tickers(data_dir)
    sources["yahoo_fetch"] = smoke
    sources_path.write_text(json.dumps(sources, indent=2))
    return data_dir, smoke


def main():
    failures = []

    with tempfile.TemporaryDirectory() as tmp:
        data_dir, smoke = _trimmed_data_dir(tmp)
        if len(smoke) < SMOKE_COUNT:
            print(f"FAIL: only {len(smoke)} eligible smoke tickers: {smoke}")
            return 1
        print(f"[live-smoke] tickers: {', '.join(smoke)}")
        output = Path(tmp) / "live_smoke.xlsx"

        # No --offline: this is the point of the check.
        build_alert_levels.main([
            "--output", str(output),
            "--data-dir", str(data_dir),
        ])

        if not output.exists():
            print("FAIL: no workbook was written")
            return 1

        from openpyxl import load_workbook

        wb = load_workbook(output)

        expected = ["Daily Summary", "Alert Levels", "Fib Methodology", "How to Run"]
        if wb.sheetnames != expected:
            failures.append(f"sheet names {wb.sheetnames} != {expected}")

        alerts = wb["Alert Levels"]

        # The banner distinguishes full/partial live from static fallback. If the
        # provider returned nothing, this is the line that says so.
        banner = alerts["A2"].value or ""
        if "live refresh" not in banner.lower():
            failures.append(f"no live data reached the workbook — banner: {banner[:90]}")

        # At least one smoke ticker must carry a real price and a real MA. A
        # provider change that empties the payload leaves these blank while the
        # run still "succeeds".
        priced = mas = 0
        for row in range(4, alerts.max_row + 1):
            if alerts.cell(row, 2).value not in smoke:
                continue
            if isinstance(alerts.cell(row, 4).value, (int, float)):
                priced += 1
            if isinstance(alerts.cell(row, 24).value, (int, float)):   # D200 SMA
                mas += 1

        if not priced:
            failures.append("no smoke ticker received a numeric current price")
        if not mas:
            failures.append("no smoke ticker received a numeric D200 SMA")

        print(
            f"[live-smoke] {priced}/{len(smoke)} priced, "
            f"{mas}/{len(smoke)} with D200"
        )

    if failures:
        print("\nFAIL:")
        for failure in failures:
            print(f"  - {failure}")
        return 1

    print("[live-smoke] OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
